import os
import pickle
import faiss
import pandas as pd
import numpy as np
from sentence_transformers import SentenceTransformer
from app_helpers import html_to_plain_text
import traceback
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
import html # Added for escaping characters

def openpyxl_rich_text_to_html(cell):
    """
    Safely converts a cell's value (including rich text with multiple formats)
    into a single HTML string.
    """
    if cell.value is None:
        return ""

    # Handle cells with rich text (multiple formatting parts)
    if cell.data_type == 'r' and isinstance(cell.value, (list, tuple)):
        html_parts = []
        for part in cell.value:
            text = getattr(part, 'text', str(part))
            font = getattr(part, 'font', None)
            
            if text is None:
                continue

            text = html.escape(str(text)).replace('\n', '<br>')
            
            styles = []
            if font:
                if font.bold:
                    styles.append('font-weight: bold;')
                if font.italic:
                    styles.append('font-style: italic;')
                
                # Safely handle font color
                if font.color and hasattr(font.color, 'rgb') and font.color.rgb:
                    rgb_value = font.color.rgb
                    if isinstance(rgb_value, str) and len(rgb_value) >= 6:
                        # Get RRGGBB part from AARRGGBB or RRGGBB
                        color_hex = rgb_value[-6:]
                        if color_hex.lower() != '000000': # Don't style default black text
                            styles.append(f'color: #{color_hex};')

            if styles:
                html_parts.append(f'<span style="{" ".join(styles)}">{text}</span>')
            else:
                html_parts.append(text)
        return "".join(html_parts)
    
    # Fallback for non-rich-text cells that may still have cell-level formatting
    text = html.escape(str(cell.value)).replace('\n', '<br>')
    styles = []
    if cell.font:
        if cell.font.bold:
            styles.append('font-weight: bold;')
        if cell.font.italic:
            styles.append('font-style: italic;')
        if cell.font.color and hasattr(cell.font.color, 'rgb') and cell.font.color.rgb:
            rgb_value = cell.font.color.rgb
            if isinstance(rgb_value, str) and len(rgb_value) >= 6:
                color_hex = rgb_value[-6:]
                if color_hex.lower() != '000000':
                    styles.append(f'color: #{color_hex};')

    if styles:
        return f'<span style="{" ".join(styles)}">{text}</span>'
    
    return text


def _process_excel_file(filepath, sentence_model, markup, is_local=False):
    """
    Processes a single Excel price list file, calculates prices, generates search text,
    creates embeddings, and returns a Faiss index and the searchable data.
    """
    if not os.path.exists(filepath):
        print(f"WARNING: Price list file not found at '{filepath}', skipping.")
        return None, None
    try:
        # Use openpyxl to load the workbook to access rich text
        wb = load_workbook(filepath, data_only=True)
        all_items = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Ensure header is read as plain strings
            header = [str(cell.value).strip() if cell.value is not None else '' for cell in ws[1]]
            
            # Find column indices
            try:
                desc_col_idx = header.index('description')
                po_price_col_idx = header.index('po_price')
            except ValueError:
                print(f"Skipping sheet '{sheet_name}' in '{filepath}' due to missing 'description' or 'po_price' column.")
                continue

            # Map other columns, defaulting to None if not found
            col_map = {col: (header.index(col) if col in header else -1) for col in ['item_code', 'make', 'approvals', 'model', 'installation', 'unit']}

            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                description_cell = row[desc_col_idx]
                # Skip empty rows
                if not description_cell.value:
                    continue
                
                po_price_cell = row[po_price_col_idx]
                po_price = pd.to_numeric(po_price_cell.value, errors='coerce')
                if pd.isna(po_price) or po_price <= 0:
                    continue

                # Get rich text description as HTML
                description_html = openpyxl_rich_text_to_html(description_cell)

                item_data = {
                    'product_type': sheet_name,
                    'description': description_html, # Store HTML description
                    'po_price': po_price,
                    'item_code': row[col_map['item_code']].value if col_map['item_code'] != -1 else (f"local_{row_idx}" if is_local else None),
                    'make': row[col_map['make']].value if col_map['make'] != -1 else None,
                    'approvals': row[col_map['approvals']].value if col_map['approvals'] != -1 else None,
                    'model': row[col_map['model']].value if col_map['model'] != -1 else None,
                    'installation': row[col_map['installation']].value if col_map['installation'] != -1 else None,
                    'unit': row[col_map['unit']].value if col_map['unit'] != -1 else 'Pcs',
                }
                all_items.append(item_data)

        if not all_items:
            return None, None

        price_list_df = pd.DataFrame(all_items)
        
        po_price_numeric = pd.to_numeric(price_list_df['po_price'], errors='coerce')
        price_list_df['offer_price'] = (po_price_numeric * (1 + markup)).round(2)
        
        # Create plain search_text from the HTML description for the search index
        price_list_df['search_text'] = price_list_df['description'].apply(html_to_plain_text).str.lower()
        
        price_list_df['is_local'] = is_local
        
        searchable_data = price_list_df.replace({np.nan: None}).to_dict('records')
        search_texts = [text for text in price_list_df['search_text'].tolist() if str(text).strip()]
        
        if search_texts:
            embeddings = sentence_model.encode(search_texts, convert_to_tensor=True).cpu().numpy()
            faiss_index = faiss.IndexFlatL2(embeddings.shape[1])
            faiss_index.add(embeddings)
            return faiss_index, searchable_data
        else:
            return None, None
            
    except Exception as e:
        print(f"Error processing Excel file {filepath}: {e}")
        traceback.print_exc()
        return None, None

# Rest of the code remains the same...
def process_and_index_data(config, sentence_model):
    """
    Orchestrates the processing of all data files (foreign items, local items, clients)
    and saves the generated indexes and data to disk.
    """
    print("Starting data processing and indexing from source files...")
    
    # Process Foreign Items
    item_faiss_index, item_searchable_data = _process_excel_file(config['PRICE_LIST_FILE'], sentence_model, config['MARKUP'], is_local=False)
    if item_faiss_index and item_searchable_data:
        faiss.write_index(item_faiss_index, config['ITEM_FAISS_INDEX_FILE'])
        with open(config['ITEM_SEARCH_DATA_FILE'], 'wb') as f:
            pickle.dump(item_searchable_data, f)
        print("Foreign item search index created.")
    else:
        print("Failed to process foreign items or no data found.")
        
    # Process Local Items
    local_item_faiss_index, local_item_searchable_data = _process_excel_file(config['LOCAL_PRICE_LIST_FILE'], sentence_model, config['MARKUP'], is_local=True)
    if local_item_faiss_index and local_item_searchable_data:
        faiss.write_index(local_item_faiss_index, config['LOCAL_ITEM_FAISS_INDEX_FILE'])
        with open(config['LOCAL_ITEM_SEARCH_DATA_FILE'], 'wb') as f:
            pickle.dump(local_item_searchable_data, f)
        print("Local item search index created.")
    else:
        print("Failed to process local items or no data found.")
        
    # Process Clients
    try:
        clients_df = pd.read_csv(config['CLIENTS_FILE'])
        clients_df['search_text'] = (clients_df['client_name'].fillna('') + ' ' + clients_df['client_address'].fillna('')).str.lower()
        client_searchable_data = clients_df.to_dict('records')
        client_embeddings = sentence_model.encode(clients_df['search_text'].tolist(), convert_to_tensor=True).cpu().numpy()
        client_faiss_index = faiss.IndexFlatL2(client_embeddings.shape[1])
        client_faiss_index.add(client_embeddings)
        
        faiss.write_index(client_faiss_index, config['CLIENT_FAISS_INDEX_FILE'])
        with open(config['CLIENT_SEARCH_DATA_FILE'], 'wb') as f:
            pickle.dump(client_searchable_data, f)
        print("Client search index created.")
    except Exception as e:
        print(f"Error processing client data: {e}")
        return False
        
    return True

def initialize_data(config, force_rebuild=False):
    """
    Initializes all data for the application. It loads the sentence transformer model,
    and then either loads the pre-built Faiss indexes from disk or rebuilds them
    by calling process_and_index_data.
    """
    # Print the model name being used to the terminal
    print(f"Initializing sentence transformer with model: {config['MODEL_NAME']}")
    sentence_model = SentenceTransformer(config['MODEL_NAME'], trust_remote_code=True)
    
    # Check if a rebuild is forced or if any of the essential index files are missing.
    rebuild_needed = force_rebuild or not all(os.path.exists(config[f]) for f in [
        'ITEM_FAISS_INDEX_FILE', 'ITEM_SEARCH_DATA_FILE',
        'LOCAL_ITEM_FAISS_INDEX_FILE', 'LOCAL_ITEM_SEARCH_DATA_FILE',
        'CLIENT_FAISS_INDEX_FILE', 'CLIENT_SEARCH_DATA_FILE'
    ])

    if rebuild_needed:
        if force_rebuild: print("Forcing a rebuild of all data indexes.")
        else: print("One or more index files are missing. Rebuilding all data.")
        success = process_and_index_data(config, sentence_model)
        if not success:
            raise Exception("Failed to build necessary data indexes.")

    # Load all data and indexes into memory
    try:
        print("Loading data indexes from disk...")
        item_faiss_index = faiss.read_index(config['ITEM_FAISS_INDEX_FILE'])
        with open(config['ITEM_SEARCH_DATA_FILE'], 'rb') as f:
            item_searchable_data = pickle.load(f)
            
        local_item_faiss_index = faiss.read_index(config['LOCAL_ITEM_FAISS_INDEX_FILE'])
        with open(config['LOCAL_ITEM_SEARCH_DATA_FILE'], 'rb') as f:
            local_item_searchable_data = pickle.load(f)

        client_faiss_index = faiss.read_index(config['CLIENT_FAISS_INDEX_FILE'])
        with open(config['CLIENT_SEARCH_DATA_FILE'], 'rb') as f:
            client_searchable_data = pickle.load(f)
            
        print(f"Indexes loaded: Foreign Items ({item_faiss_index.ntotal}), Local Items ({local_item_faiss_index.ntotal}), Clients ({client_faiss_index.ntotal}).")
        
        return {
            "sentence_model": sentence_model,
            "item_faiss_index": item_faiss_index,
            "item_searchable_data": item_searchable_data,
            "local_item_faiss_index": local_item_faiss_index,
            "local_item_searchable_data": local_item_searchable_data,
            "client_faiss_index": client_faiss_index,
            "client_searchable_data": client_searchable_data,
        }
        
    except Exception as e:
        print(f"A critical error occurred while loading data indexes: {e}")
        traceback.print_exc()
        raise