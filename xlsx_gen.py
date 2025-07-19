import os
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.cell.rich_text import TextBlock, CellRichText, InlineFont
from app_helpers import to_words_usd, to_words_bdt

# --- Helper to prevent conversion errors ---
def safe_float(value, default=0.0):
    try:
        if value is None or value == '':
            return default
        return float(value)
    except (ValueError, TypeError):
        return default

# --- XLSX Generation Functions ---
def parse_html_to_richtext(html_string):
    if not isinstance(html_string, str): return str(html_string) if html_string is not None else ''
    html_string = str(html_string).replace('&nbsp;', ' ')
    html_string = re.sub(r'<br\s*/?>', '\n', html_string)
    parts = re.split(r'(<.*?>)', html_string)
    text_blocks, is_bold, is_italic, color = [], False, False, "000000"

    for part in parts:
        if not part: continue
        is_tag = part.startswith('<') and part.endswith('>')
        if is_tag:
            tag_match = re.match(r'</?([a-zA-Z0-9]+)', part)
            if not tag_match: continue
            tag = tag_match.group(1).lower()

            is_opening_tag = not part.startswith('</')
            if tag == 'b': is_bold = is_opening_tag
            elif tag == 'i': is_italic = is_opening_tag
            elif tag == 'font':
                if is_opening_tag:
                    color_match = re.search(r'color="#?([0-9a-fA-F]{6})"', part)
                    if color_match: color = color_match.group(1)
                else: color = "000000"
        else:
            text = part.replace('&lt;', '<').replace('&gt;', '>').replace('&amp;', '&')
            font_props = InlineFont()
            if is_bold:
                font_props.b = True
            if is_italic:
                font_props.i = True
            if color != "000000":
                font_props.color = color
            text_blocks.append(TextBlock(font_props, text))
    if not text_blocks:
        return CellRichText([TextBlock(InlineFont(), '')])
    return CellRichText(text_blocks)

def add_tnc_to_excel(wb, tnc_text, header_color_hex, auth_dir, data):
    if not tnc_text: return
    ws = wb.create_sheet("Terms & Conditions")
    bold_font, header_font = Font(bold=True, size=12), Font(bold=True)
    highlight_fill = PatternFill(start_color=header_color_hex, end_color=header_color_hex, fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.column_dimensions['A'].width, ws.column_dimensions['B'].width, ws.column_dimensions['C'].width = 5, 80, 30
    row_cursor = 1

    title_cell = ws.cell(row=row_cursor, column=1, value="Terms & Conditions")
    title_cell.font = Font(bold=True, size=16)
    title_cell.fill = highlight_fill

    ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=3)
    ws.cell(row=row_cursor, column=1).alignment = Alignment(horizontal='center'); row_cursor += 2
    in_table, table_data = False, []
    tnc_headers = ["Foreign Part:", "Local Part (Supply):", "Local Part (Installation):"]
    
    headers_needing_space = [
        "Local Part (Supply):",
        "Payment Schedule (Local Supply):",
        "Local Part (Installation):",
        "Payment Schedule (Installation):"
    ]


    for line in tnc_text.split('\n'):
        line = line.strip()
        if not line: continue

        if any(h in line for h in headers_needing_space):
            row_cursor += 1

        if line == 'TABLE_START': in_table, table_data = True, []; continue
        elif line == 'TABLE_END':
            in_table = False
            if table_data:
                headers = [h.strip() for h in table_data[0].split('|')]
                col_widths = [10, 80, 25]
                for c_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_cursor, column=c_idx, value=header)
                    cell.font, cell.fill, cell.border, cell.alignment = header_font, highlight_fill, thin_border, Alignment(horizontal='center', vertical='center', wrap_text=True)
                    if c_idx -1 < len(col_widths): ws.column_dimensions[chr(64 + c_idx)].width = col_widths[c_idx-1]

                row_cursor += 1
                for row_line in table_data[1:]:
                    row = [r.strip() for r in row_line.split('|')]
                    for c_idx, cell_text in enumerate(row, 1):
                        cell = ws.cell(row=row_cursor, column=c_idx, value=cell_text)
                        cell.border = thin_border
                        if c_idx == 1 or c_idx == 3:
                            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                        else:
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                    row_cursor += 1
            row_cursor += 1; continue
        if in_table: table_data.append(line); continue

        if any(header in line for header in tnc_headers):
            cell = ws.cell(row=row_cursor, column=1, value=line)
            cell.font = bold_font
            cell.fill = highlight_fill
            ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=3)
            row_cursor += 1
        elif re.match(r'^[A-Za-z]+\s*Part.*:$', line) or "Payment Schedule" in line:
            cell = ws.cell(row=row_cursor, column=1, value=line)
            cell.font = bold_font
            cell.fill = highlight_fill
            ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=3); row_cursor += 1
        else:
            ws.cell(row=row_cursor, column=1, value=line)
            ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=3)
            ws.cell(row=row_cursor, column=1).alignment = Alignment(wrap_text=True, vertical='top')
            row_cursor += 1

    row_cursor += 3
    ws.cell(row=row_cursor, column=1, value="Sincerely,").font = bold_font
    row_cursor += 1

    if data.get('includeSignature', True):
        signature_image_path = os.path.join(auth_dir, 'Signature_RIF.png')
        if os.path.exists(signature_image_path):
            img = Image(signature_image_path)

            aspect_ratio = img.width / img.height
            target_height = 118
            img.height = target_height
            img.width = target_height * aspect_ratio

            ws.add_image(img, f'A{row_cursor}')
    row_cursor += 6

    ws.cell(row=row_cursor, column=1, value="Md. Rezwanul Islam").font = bold_font; row_cursor += 1
    ws.cell(row=row_cursor, column=1, value="Manager, Business Development"); row_cursor += 1
    ws.cell(row=row_cursor, column=1, value="AMO Green Energy Limited"); row_cursor += 1

def draw_financial_summary_for_boq(ws, data, visible_price_sections, header_row_2_num, financial_labels, is_local_only=False):
    financials = data.get('financials', {})
    items = data.get('items', [])
    has_additional_charges = data.get('has_additional_charges', False)

    header_fill = PatternFill(start_color="EEE576", end_color="EEE576", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    bold_font = Font(bold=True)
    grand_total_font = Font(bold=True, size=12)
    red_bold_font = Font(bold=True, color="FF0000")
    right_align = Alignment(horizontal='right', vertical='center')

    value_cols = {}
    current_price_col_start = 5 
    for section in visible_price_sections:
        value_cols[section['key']] = current_price_col_start + 1
        current_price_col_start += 2

    full_header_count = ws.max_column
    ws.append([])
    current_row = ws.max_row

    subtotals = {}
    for section in visible_price_sections:
        total_key = section['sub'][1]['key']
        subtotals[section['key']] = sum(safe_float(it.get(total_key, 0)) for it in items)
    
    freight = safe_float(financials.get('freight_foreign_usd')) if financials.get('use_freight') else 0
    delivery = safe_float(financials.get('delivery_local_bdt')) if financials.get('use_delivery') else 0
    vat = safe_float(financials.get('vat_local_bdt')) if financials.get('use_vat') else 0
    ait = safe_float(financials.get('ait_local_bdt')) if financials.get('use_ait') else 0
    discount_foreign = safe_float(financials.get('discount_foreign_usd')) if financials.get('use_discount_foreign') else 0
    discount_local = safe_float(financials.get('discount_local_bdt')) if financials.get('use_discount_local') else 0
    discount_install = safe_float(financials.get('discount_installation_bdt')) if financials.get('use_discount_installation') else 0

    grand_total_foreign = subtotals.get('foreign', 0) + freight - discount_foreign
    grand_total_local_supply = subtotals.get('local_supply_price', 0) - discount_local
    grand_total_install = subtotals.get('installation_price', 0) - discount_install

    visible_section_keys = [s['key'] for s in visible_price_sections]
    if 'local_supply_price' in visible_section_keys:
        grand_total_local_supply += delivery + vat + ait
    elif 'installation_price' in visible_section_keys:
        grand_total_install += delivery + vat + ait
        
    grand_totals = {
        'foreign': grand_total_foreign,
        'local_supply_price': grand_total_local_supply,
        'installation_price': grand_total_install
    }
    
    grand_total_label = financial_labels.get('grandtotalLocal', 'Grand Total (BDT):') if is_local_only else financial_labels.get('grandtotalForeign', 'Grand Total:')

    first_val_col_idx = 5
    if value_cols:
        first_val_col_idx = min(v for v in value_cols.values() if v is not None)
    label_col_end_idx = first_val_col_idx - 1

    def add_financial_row(label, values_dict, is_bold=False, is_red=False, is_grand_total=False):
        nonlocal current_row
        ws.append([])
        current_row += 1
        if label_col_end_idx >= 1:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=label_col_end_idx)

        label_cell = ws.cell(row=current_row, column=1, value=label)
        label_cell.alignment = right_align
        if is_bold or is_grand_total:
            label_cell.font = grand_total_font if is_grand_total else bold_font
        if is_red:
            label_cell.font = red_bold_font

        for key, value in values_dict.items():
            col_idx = value_cols.get(key)
            if not col_idx or value is None: continue
            
            value_cell = ws.cell(row=current_row, column=col_idx, value=value)
            
            price_header_cell = ws.cell(row=header_row_2_num, column=col_idx-1)
            
            if 'USD' in price_header_cell.value: value_cell.number_format = '"$"#,##0.00'
            elif 'BDT' in price_header_cell.value: value_cell.number_format = '"BDT "#,##0.00'
            
            if is_bold or is_grand_total:
                value_cell.font = grand_total_font if is_grand_total else bold_font
            if is_red:
                value_cell.font = red_bold_font
            value_cell.alignment = right_align

        for col in range(1, full_header_count + 1):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            if is_grand_total:
                 cell.fill = header_fill

    if has_additional_charges:
        # NEW: Added separate sub-total for local supply
        local_supply_subtotal = {'local_supply_price': subtotals.get('local_supply_price', 0)}
        if local_supply_subtotal['local_supply_price'] > 0:
            add_financial_row("Sub-Total (Local Supply):", local_supply_subtotal, is_bold=True)
            
        if 'installation_price' in visible_section_keys:
            install_subtotal = {'installation_price': subtotals.get('installation_price', 0)}
            add_financial_row("Sub-Total (Installation):", install_subtotal, is_bold=True)

        is_foreign_visible_boq = 'foreign' in visible_section_keys
        is_local_supply_visible_boq = 'local_supply_price' in visible_section_keys
        is_install_visible_boq = 'installation_price' in visible_section_keys

        if freight > 0 and is_foreign_visible_boq:
            add_financial_row(financial_labels.get('freight', 'Freight:'), {'foreign': freight})
        
        if delivery > 0 and (is_local_supply_visible_boq or is_install_visible_boq):
            add_financial_row(financial_labels.get('delivery', 'Delivery:'), {'local_supply_price' if is_local_supply_visible_boq else 'installation_price': delivery})
        if vat > 0 and (is_local_supply_visible_boq or is_install_visible_boq):
            add_financial_row(financial_labels.get('vat', 'VAT:'), {'local_supply_price' if is_local_supply_visible_boq else 'installation_price': vat})
        if ait > 0 and (is_local_supply_visible_boq or is_install_visible_boq):
            add_financial_row(financial_labels.get('ait', 'AIT:'), {'local_supply_price' if is_local_supply_visible_boq else 'installation_price': ait})
            
        if any(d > 0 for d in [discount_foreign, discount_local, discount_install]):
            discount_values = {
                'foreign': -discount_foreign if discount_foreign > 0 else None,
                'local_supply_price': -discount_local if discount_local > 0 else None,
                'installation_price': -discount_install if discount_install > 0 else None
            }
            add_financial_row("Discount:", {k: v for k, v in discount_values.items() if v is not None}, is_red=True, is_bold=True)
    
    add_financial_row(grand_total_label, grand_totals, is_grand_total=True)


def draw_boq_sheet(ws, data, header_color_hex, financial_labels, is_local_only):
    user_info, items = data.get('user', {}), data.get('items', [])
    visible_columns = data.get('visibleColumns', {})
    user_role = user_info.get('role', 'user')

    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color=header_color_hex, end_color=header_color_hex, fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')
    left_align_top_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    is_admin = user_role == 'admin'

    price_sections = [
        {'key': 'foreign', 'header': financial_labels.get('foreignPrice', 'FOREIGN PRICE'), 'visible': visible_columns.get('foreign_price', True), 'sub': [
            {'header': 'PRICE\n(USD)', 'key': 'foreign_price_usd'},
            {'header': 'TOTAL\n(USD)', 'key': 'foreign_total_usd'}
        ]},
        {'key': 'po_price', 'header': 'PO PRICE', 'visible': is_admin and visible_columns.get('po_price'), 'sub': [
            {'header': 'PRICE\n(USD)', 'key': 'po_price_usd'},
            {'header': 'TOTAL\n(USD)', 'key': 'po_total_usd'}
        ]},
        {'key': 'local_supply_price', 'header': financial_labels.get('localPrice', 'LOCAL SUPPLY PRICE'), 'visible': visible_columns.get('local_supply_price'), 'sub': [
            {'header': 'PRICE\n(BDT)', 'key': 'local_supply_price_bdt'},
            {'header': 'TOTAL\n(BDT)', 'key': 'local_supply_total_bdt'}
        ]},
        {'key': 'installation_price', 'header': financial_labels.get('installationPrice', 'INSTALLATION PRICE'), 'visible': visible_columns.get('installation_price'), 'sub': [
            {'header': 'PRICE\n(BDT)', 'key': 'installation_price_bdt'},
            {'header': 'TOTAL\n(BDT)', 'key': 'installation_total_bdt'}
        ]},
    ]
    visible_price_sections = [s for s in price_sections if s['visible']]
    
    base_headers = ['SL', 'Description', 'Qty', 'Unit']
    
    header1_data = base_headers + [None] * (len(visible_price_sections) * 2)
    header2_data = [None] * len(base_headers) + [sub['header'] for section in visible_price_sections for sub in section['sub']]
    ws.append(header1_data)
    ws.append(header2_data)

    header_row_1_num = ws.max_row - 1
    header_row_2_num = ws.max_row
    
    for col_idx in range(1, len(base_headers) + 1):
        ws.merge_cells(start_row=header_row_1_num, start_column=col_idx, end_row=header_row_2_num, end_column=col_idx)

    current_col = len(base_headers) + 1
    for section in visible_price_sections:
        ws.merge_cells(start_row=header_row_1_num, start_column=current_col, end_row=header_row_1_num, end_column=current_col + 1)
        main_header_cell = ws.cell(row=header_row_1_num, column=current_col)
        main_header_cell.value = section['header']
        current_col += 2
        
    for row in ws.iter_rows(min_row=header_row_1_num, max_row=header_row_2_num, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
    
    for i, item in enumerate(items):
        row_data = [i + 1, parse_html_to_richtext(item.get('description', '')), int(item.get('qty', 1)), item.get('unit', 'Pcs')]
        row_data.extend([safe_float(item.get(sub['key'], 0)) for section in visible_price_sections for sub in section['sub']])
        
        ws.append(row_data)
        current_row_obj = ws[ws.max_row]
        for idx, cell in enumerate(current_row_obj):
            cell.border = thin_border
            if idx == 1: cell.alignment = left_align_top_wrap
            elif idx in [0, 2, 3]: cell.alignment = center_align
            else:
                cell.alignment = right_align
                header_text = ws.cell(row=header_row_2_num, column=idx + 1).value
                if 'USD' in header_text:
                    cell.number_format = '"$"#,##0.00'
                elif 'BDT' in header_text:
                    cell.number_format = '"BDT "#,##0.00'
                else:
                    cell.number_format = '#,##0.00'

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 10
    col_letter_start_idx = 5
    for section in visible_price_sections:
        for _ in section['sub']:
            ws.column_dimensions[chr(64 + col_letter_start_idx)].width = 18
            col_letter_start_idx += 1

    draw_financial_summary_for_boq(ws, data, visible_price_sections, header_row_2_num, financial_labels, is_local_only)

def generate_financial_offer_xlsx(data, auth_dir, header_color_hex):
    wb = Workbook()
    
    items = data.get('items', [])
    financials = data.get('financials', {})

    subtotal_foreign = sum(safe_float(it.get('foreign_total_usd', 0)) for it in items)
    subtotal_local_supply = sum(safe_float(it.get('local_supply_total_bdt', 0)) for it in items)
    subtotal_installation = sum(safe_float(it.get('installation_total_bdt', 0)) for it in items)

    freight = safe_float(financials.get('freight_foreign_usd')) if financials.get('use_freight') else 0
    delivery = safe_float(financials.get('delivery_local_bdt')) if financials.get('use_delivery') else 0
    vat = safe_float(financials.get('vat_local_bdt')) if financials.get('use_vat') else 0
    ait = safe_float(financials.get('ait_local_bdt')) if financials.get('use_ait') else 0
    discount_foreign = safe_float(financials.get('discount_foreign_usd')) if financials.get('use_discount_foreign') else 0
    discount_local = safe_float(financials.get('discount_local_bdt')) if financials.get('use_discount_local') else 0
    discount_install = safe_float(financials.get('discount_installation_bdt')) if financials.get('use_discount_installation') else 0

    grand_total_usd = subtotal_foreign + freight - discount_foreign
    grand_total_bdt = subtotal_local_supply + subtotal_installation + delivery + vat + ait - discount_local - discount_install
    
    data['words_usd'] = to_words_usd(grand_total_usd)
    data['words_bdt'] = to_words_bdt(grand_total_bdt)
    
    is_summary_enabled = data.get('isSummaryPageEnabled', False)
    financial_labels = data.get('financialLabels', {})
    
    client_info = data.get('client', {})
    full_reference_number = data.get('referenceNumber', "FinancialOffer_NoRef")
    display_reference = full_reference_number.split('_')[-1] if '_' in full_reference_number else full_reference_number
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color=header_color_hex, end_color=header_color_hex, fill_type="solid")
    center_align_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    user_info = data.get('user', {})
    visible_columns, user_role = data.get('visibleColumns', {}), user_info.get('role', 'user')
    is_foreign_visible = visible_columns.get('foreign_price', True)
    is_local_visible = visible_columns.get('local_supply_price')
    is_install_visible = visible_columns.get('installation_price')
    is_local_only = not is_foreign_visible and (is_local_visible or is_install_visible)
    is_local_part_visible = is_local_visible or is_install_visible

    has_freight_and_is_visible = freight > 0 and is_foreign_visible
    if has_freight_and_is_visible:
        financial_labels['subtotalForeign'] = 'Subtotal, Ex-Works:'
        financial_labels['grandtotalForeign'] = 'Grand Total, CFR, Chattogram (USD):'
    else:
        financial_labels['subtotalForeign'] = 'Subtotal:'
        financial_labels['grandtotalForeign'] = 'Grand Total, Ex-Works (USD):'

    visible_count = sum([visible_columns.get(key) for key in ['foreign_price', 'local_supply_price', 'installation_price', 'po_price'] if visible_columns.get(key)])
    end_col = 4 + (visible_count * 2)

    if is_summary_enabled:
        ws = wb.active
        ws.title = "Financial Summary"
        summary_scopes = data.get('summaryScopes', {})
        
        title_font = Font(name='Calibri', size=14, bold=True, underline='single')
        header_font = Font(name='Calibri', size=11, bold=True)
        red_bold_font = Font(bold=True, color="FF0000")
        right_align = Alignment(horizontal='right', vertical='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        ws.merge_cells('A1:D1')
        ws['A1'] = "SUMMARY OF SUPPLY & INSTALLATION OF FIRE PROTECTION SYSTEM AND FIRE DETECTION & ALARM SYSTEM"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align_wrap
        ws['A1'].fill = header_fill

        ws.append([])
        ws.append(['Project Name:', client_info.get('name', 'N/A')])
        ws['A3'].font = bold_font
        ws.append(['Submitted By:', 'AMO Green Energy Limited'])
        ws['A4'].font = bold_font
        ws.append(['Submission Date:', datetime.now().strftime('%d-%B-%Y')])
        ws['A5'].font = bold_font
        ws.append(['Reference:', display_reference])
        ws['A6'].font = bold_font
        ws.append([])

        ws.merge_cells('A8:D8')
        ws['A8'] = "PRICE SUMMARY"
        ws['A8'].font = Font(size=12, bold=True)
        ws['A8'].alignment = center_align_wrap
        ws['A8'].fill = header_fill
        ws.append([])

        ws.merge_cells('A10:A11')
        ws['A10'] = "SL"
        ws.merge_cells('B10:B11')
        ws['B10'] = "Scope of Works"
        ws.merge_cells('C10:D10')
        ws['C10'] = "Description of Head of Cost"
        ws['C11'] = "Imported Items with Sea Freight (USD)"
        ws['D11'] = "Supply, Installation, Testing & Commissioning (BDT)"

        for row in ws['A10:D11']:
            for cell in row:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = center_align_wrap
        ws.row_dimensions[11].height = 40
        
        sub_total_usd = sum(scope['total_usd'] for scope in summary_scopes.values())
        sub_total_bdt = sum(scope['total_bdt'] for scope in summary_scopes.values())
        
        # --- SORTING FIX START ---
        scope_order = ['foreign', 'localsupply', 'installation']
        def sort_key(item):
            key_parts = item[0].split('-')
            if len(key_parts) == 2:
                scope_type, sub_type = key_parts
                if sub_type in scope_order:
                    return (scope_order.index(sub_type), scope_type)
            return (len(scope_order), item[0])

        sorted_scopes = sorted(summary_scopes.items(), key=sort_key)
        # --- SORTING FIX END ---

        current_row_idx = 12
        for i, (key, scope) in enumerate(sorted_scopes):
            sl_cell = ws.cell(row=current_row_idx, column=1, value=chr(65 + i))
            sl_cell.alignment = center_align_wrap
            ws.cell(row=current_row_idx, column=2, value=scope['description'])
            usd_cell = ws.cell(row=current_row_idx, column=3, value=scope['total_usd'])
            bdt_cell = ws.cell(row=current_row_idx, column=4, value=scope['total_bdt'])
            usd_cell.number_format = '"$"#,##0.00'
            bdt_cell.number_format = '"BDT "#,##0.00'
            
            for col in range(1, 5):
                ws.cell(row=current_row_idx, column=col).border = thin_border
                
            current_row_idx +=1

        has_additional_charges = data.get('has_additional_charges', False)
        
        summary_data = []
        if has_additional_charges:
            summary_data.append((financial_labels.get('subtotalForeign', 'Sub-Total:'), sub_total_usd, sub_total_bdt))
            if freight > 0 and is_foreign_visible:
                summary_data.append((financial_labels.get('freight', 'Sea Freight:'), freight, None))
            if is_local_part_visible:
                if delivery > 0:
                    summary_data.append((financial_labels.get('delivery', 'Delivery Charge:'), None, delivery))
                if vat > 0:
                    summary_data.append((financial_labels.get('vat', 'VAT:'), None, vat))
                if ait > 0:
                    summary_data.append((financial_labels.get('ait', 'AIT:'), None, ait))
            if discount_foreign > 0 or (discount_local + discount_install) > 0:
                summary_data.append(("Special Discount:", -discount_foreign, -(discount_local + discount_install)))
        
        grand_total_label = financial_labels.get('grandtotalLocal', 'Grand Total (BDT):') if is_local_only else financial_labels.get('grandtotalForeign', 'Grand Total:')
        summary_data.append((grand_total_label, grand_total_usd, grand_total_bdt))


        for label, val_usd, val_bdt in summary_data:
            label_cell = ws.cell(row=current_row_idx, column=1, value=label)
            usd_cell = ws.cell(row=current_row_idx, column=3, value=val_usd if val_usd is not None else "")
            bdt_cell = ws.cell(row=current_row_idx, column=4, value=val_bdt if val_bdt is not None else "")
            
            ws.merge_cells(start_row=current_row_idx, start_column=1, end_row=current_row_idx, end_column=2)
            label_cell.alignment = right_align
            
            usd_cell.number_format = '"$"#,##0.00'
            bdt_cell.number_format = '"BDT "#,##0.00'

            is_total_row = "Total" in label
            is_discount_row = "Discount" in label
            
            if is_total_row:
                label_cell.font = bold_font
                usd_cell.font = bold_font
                bdt_cell.font = bold_font
            
            if is_discount_row:
                label_cell.font = red_bold_font
                usd_cell.font = red_bold_font
                bdt_cell.font = red_bold_font
            
            if "Grand Total" in label:
                for col in range(1, 5): ws.cell(row=current_row_idx, column=col).fill = header_fill
            
            for col in range(1, 5): ws.cell(row=current_row_idx, column=col).border = thin_border
            
            current_row_idx += 1

        current_row_idx += 1
        
        # --- "IN WORDS" FIX START ---
        if data.get('has_foreign_part'):
            in_words_text_usd = f"In Words (Foreign Part):   {data.get('words_usd', 'N/A')}"
            cell = ws.cell(row=current_row_idx, column=1, value=in_words_text_usd)
            ws.merge_cells(start_row=current_row_idx, start_column=1, end_row=current_row_idx, end_column=4)
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = center_align_wrap
            for col_idx in range(1, 5): 
                ws.cell(row=current_row_idx, column=col_idx).border = thin_border
            current_row_idx += 1

        if data.get('has_local_part'):
            in_words_text_bdt = f"In Words (Local Part):   {data.get('words_bdt', 'N/A')}"
            cell = ws.cell(row=current_row_idx, column=1, value=in_words_text_bdt)
            ws.merge_cells(start_row=current_row_idx, start_column=1, end_row=current_row_idx, end_column=4)
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = center_align_wrap
            for col_idx in range(1, 5): 
                ws.cell(row=current_row_idx, column=col_idx).border = thin_border
            current_row_idx += 1
        # --- "IN WORDS" FIX END ---
        
        if data.get('has_foreign_part') and freight > 0:
            current_row_idx += 1
            ws.cell(current_row_idx, 1, value="Important Note:").font = bold_font
            ws.cell(current_row_idx, 1).fill = header_fill
            ws.merge_cells(start_row=current_row_idx, start_column=1, end_row=current_row_idx, end_column=4)
            ws.cell(current_row_idx, 1).alignment = center_align_wrap
            
            current_row_idx += 1
            note_text = (f"Freight Charge (USD {freight:,.2f}) has been considered in light of the current market trending price, "
                         "applicable for the Sea Freight. Any increase in the freight charge will have to be borne by the client, "
                         "at the time of shipment.")
            ws.cell(row=current_row_idx, column=1, value=note_text)
            ws.merge_cells(start_row=current_row_idx, start_column=1, end_row=current_row_idx, end_column=4)
            ws.cell(current_row_idx, 1).alignment = Alignment(wrap_text=True)

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 35

        boq_ws = wb.create_sheet("Bill of Quantities")
        
        boq_title_cell = boq_ws.cell(row=1, column=1, value="Bill of Quantities")
        boq_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
        boq_title_cell.font = Font(name='Calibri', size=16, bold=True)
        boq_title_cell.fill = header_fill
        boq_title_cell.alignment = center_align_wrap
        boq_ws.append([])

        draw_boq_sheet(boq_ws, data, header_color_hex, financial_labels, is_local_only)
        
    else:
        ws = wb.active
        ws.title = "Financial Offer"
        
        title_cell = ws.cell(row=1, column=1, value="Financial Offer")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
        title_cell.font = Font(name='Calibri', size=16, bold=True)
        title_cell.fill = header_fill
        title_cell.alignment = center_align_wrap
        ws.append([])
        
        ws.append(['Ref:', display_reference])
        ws['A3'].font = bold_font
        ws.append(['Client:', client_info.get('name', 'N/A')])
        ws['A4'].font = bold_font
        ws.append(['Address:', client_info.get('address', 'N/A')])
        ws['A5'].font = bold_font
        ws.append([])
        
        draw_boq_sheet(ws, data, header_color_hex, financial_labels, is_local_only)
        
        current_row_idx = ws.max_row + 2
        
        max_col_for_merge = ws.max_column

        # --- "IN WORDS" FIX START ---
        if data.get('has_foreign_part'):
            in_words_text = f"In Words (Foreign Part):   {data.get('words_usd', 'N/A')}"
            cell = ws.cell(row=current_row_idx, column=1, value=in_words_text)
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = center_align_wrap
            ws.merge_cells(start_row=current_row_idx, start_column=1, end_row=current_row_idx, end_column=max_col_for_merge)
            current_row_idx += 1

        if data.get('has_local_part'):
            in_words_text = f"In Words (Local Part):   {data.get('words_bdt', 'N/A')}"
            cell = ws.cell(row=current_row_idx, column=1, value=in_words_text)
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = center_align_wrap
            ws.merge_cells(start_row=current_row_idx, start_column=1, end_row=current_row_idx, end_column=max_col_for_merge)
            current_row_idx += 1
        # --- "IN WORDS" FIX END ---
        
        if data.get('has_foreign_part') and freight > 0:
            current_row_idx += 1
            ws.cell(current_row_idx, 1, value="Important Note:").font = bold_font
            ws.cell(current_row_idx, 1).fill = header_fill
            ws.merge_cells(start_row=current_row_idx, start_column=1, end_row=current_row_idx, end_column=ws.max_column)
            ws.cell(current_row_idx, 1).alignment = Alignment(horizontal='center', vertical='center')
            
            current_row_idx += 1
            note_text = (f"Freight Charge (USD {freight:,.2f}) has been considered in light of the current market trending price, "
                         "applicable for the Sea Freight. Any increase in the freight charge will have to be borne by the client, "
                         "at the time of shipment.")
            ws.cell(row=current_row_idx, column=1, value=note_text)
            ws.merge_cells(start_row=current_row_idx, start_column=1, end_row=current_row_idx, end_column=ws.max_column)
            ws.cell(current_row_idx, 1).alignment = Alignment(wrap_text=True)

    add_tnc_to_excel(wb, data.get('terms_and_conditions', ''), header_color_hex, auth_dir, data)
    
    if 'Financial Summary' in wb.sheetnames:
        wb.active = wb['Financial Summary']
    elif 'Financial Offer' in wb.sheetnames:
        wb.active = wb['Financial Offer']
        
    return wb


def generate_purchase_order_xlsx(po_data, auth_dir, header_color_hex="D6EAF8"):
    items, project_info, financials = po_data.get('items', []), po_data.get('project_info', {}), po_data.get('financials', {})
    original_fo_ref = project_info.get('referenceNumber', 'UnknownRef')
    po_reference_number = original_fo_ref.replace('FO_', 'PO_', 1) if 'FO_' in original_fo_ref else f"PO_{original_fo_ref}"
    
    wb = Workbook(); ws = wb.active; ws.title = "Purchase Order"
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color=header_color_hex, end_color=header_color_hex, fill_type="solid")
    center_align, right_align, left_align = Alignment(horizontal='center', vertical='center', wrap_text=True), Alignment(horizontal='right'), Alignment(horizontal='left', vertical='top', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    grand_total_font, bold_font = Font(bold=True, size=12), Font(bold=True)
    ws.merge_cells('A1:F1'); ws['A1'] = 'Purchase Order'; ws['A1'].font = Font(bold=True, size=18); ws['A1'].alignment = center_align; ws['A1'].fill = header_fill
    ws.append([])
    ws.append([f"PO Reference:", po_reference_number])
    ws.append([f"For Project/Client:", project_info.get('client', {}).get('name', 'N/A')])
    ws.append([f"Original FO Ref:", original_fo_ref]); ws.append([])
    headers = ['SL', 'Description', 'PO Price', 'Qty', 'Unit', 'Total']
    ws.append(headers)
    header_row = ws[ws.max_row]
    for cell in header_row: cell.font, cell.fill, cell.alignment, cell.border = header_font, header_fill, center_align, thin_border

    grand_total = 0
    for i, item in enumerate(items):
        po_total = safe_float(item.get('po_total_usd', 0))
        grand_total += po_total
        row_data = [i + 1, parse_html_to_richtext(item.get('description', '')), safe_float(item.get('po_price_usd', 0)), int(item.get('qty', 1)), item.get('unit', 'Pcs'), po_total]
        ws.append(row_data)
        current_row = ws[ws.max_row]
        for idx, cell in enumerate(current_row):
            cell.border = thin_border
            if idx == 1: cell.alignment = left_align
            else: cell.alignment = right_align
            if isinstance(cell.value, (int, float)): cell.number_format = '#,##0.00'
    
    ws.append([])
    ws.cell(row=ws.max_row + 1, column=len(headers) - 1, value="Grand Total:").font = grand_total_font
    total_cell = ws.cell(row=ws.max_row, column=len(headers), value=grand_total)
    total_cell.font = grand_total_font
    total_cell.number_format = '#,##0.00'
    ws.cell(row=ws.max_row -1, column=len(headers)-1).alignment = right_align
    ws.cell(row=ws.max_row -1, column=len(headers)).alignment = right_align


    for col, width in {'A': 5, 'B': 60, 'C': 15, 'D': 10, 'E': 10, 'F': 20}.items(): ws.column_dimensions[col].width = width
    return wb

def generate_challan_xlsx(data, auth_dir, header_color_hex):
    client_info = data.get('client', {})
    items = data.get('items', [])
    ref_number = data.get('referenceNumber', 'N/A')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Delivery Challan"
    
    header_font = Font(bold=True, size=18)
    table_header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color=header_color_hex, end_color=header_color_hex, fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    left_align_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
    center_align_top = Alignment(horizontal='center', vertical='top')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))


    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = 'DELIVERY CHALLAN'
    title_cell.font = header_font
    title_cell.alignment = center_align
    title_cell.fill = header_fill
    
    ws.append([])
    ws.append([f"Challan No:", ref_number])
    challan_row_num = ws.max_row
    ws.cell(row=challan_row_num, column=1).font = table_header_font
    ws.cell(row=challan_row_num, column=2).alignment = left_align
    ws.append([f"Client:", client_info.get('name', 'N/A')])
    ws.cell(row=ws.max_row, column=1).font = table_header_font
    ws.append([f"Address:", client_info.get('address', 'N/A')])
    ws.cell(row=ws.max_row, column=1).font = table_header_font
    ws.append([])
    
    headers = ['SL', 'Item Description', 'Quantity', 'Unit']
    ws.append(headers)
    header_row_obj = ws[ws.max_row]
    for cell in header_row_obj:
        cell.font = table_header_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border


    for i, item in enumerate(items):
        quantity_val = int(item.get('qty', 1))
        unit_val = item.get('unit', 'Pcs')
        row_data = [i + 1, parse_html_to_richtext(item.get('description', '')), quantity_val, unit_val]
        ws.append(row_data)
        current_row_obj = ws[ws.max_row]
        current_row_obj[0].alignment = center_align_top
        current_row_obj[0].border = thin_border
        current_row_obj[1].alignment = left_align_wrap
        current_row_obj[1].border = thin_border
        current_row_obj[2].alignment = center_align_top
        current_row_obj[2].border = thin_border
        current_row_obj[3].alignment = center_align_top
        current_row_obj[3].border = thin_border
        ws.row_dimensions[ws.max_row].height = None

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 70
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    current_row = ws.max_row + 7

    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=4)
    
    line_cell_received = ws.cell(row=current_row, column=1, value="____________________")
    line_cell_received.font = table_header_font
    line_cell_authorized = ws.cell(row=current_row, column=3, value="____________________")
    line_cell_authorized.font = table_header_font
    line_cell_authorized.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(start_row=current_row + 1, start_column=3, end_row=current_row + 1, end_column=4)
    ws.cell(row=current_row + 1, column=1, value="Received By").font = table_header_font
    ws.cell(row=current_row + 1, column=3, value="Authorized Signature").font = table_header_font
    
    ws.cell(row=current_row + 1, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=current_row + 1, column=3).alignment = Alignment(horizontal='center')

    if data.get('includeSignature', True):
        signature_image_path = os.path.join(auth_dir, 'Signature_RIF.png')
        if os.path.exists(signature_image_path):
            img = Image(signature_image_path)
            
            aspect_ratio = img.width / img.height
            target_height = 99
            img.height = target_height
            img.width = target_height * aspect_ratio
            
            img.anchor = f'C{current_row - 4}'
            img.left = ws.column_dimensions['C'].width * 9.5
            ws.add_image(img)
    
    return wb