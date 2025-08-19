# app.py
from flask import Flask, request, jsonify, send_file, render_template, send_from_directory
from flask_socketio import SocketIO, emit
from werkzeug.utils import safe_join, secure_filename
from werkzeug.datastructures import FileStorage
from flask_cors import CORS
import pandas as pd
import numpy as np
import os
import io
from datetime import datetime
import random
import re
import csv
import json
import uuid
import math
import tnc
from openpyxl.reader.excel import load_workbook
from dotenv import load_dotenv

# Import newly created modules
import pdf_gen
import xlsx_gen
import data_management
import cover_merger
from app_helpers import html_to_plain_text, to_words_usd, to_words_bdt

# NEW: Import BeautifulSoup for HTML cleaning
from bs4 import BeautifulSoup

# Load environment variables from .env file
load_dotenv()

# Initialize Flask App
app = Flask(__name__)
CORS(app)
# Add SocketIO for real-time chat
socketio = SocketIO(app, cors_allowed_origins="*")


# --- Configuration ---
CONFIG = {
    'DATA_DIR': 'data_storage',
    'FOS_DIR': os.path.join('data_storage', 'FOS'),
    'PROJECTS_DIR': os.path.join('data_storage', 'projects'),
    'AUTH_DIR': 'authorization',
    'ASSETS_DIR': 'assets',
    'COVERS_DIR': os.path.join('assets', 'covers'),
    'THUMBNAILS_DIR': os.path.join('assets', 'covers', 'thumbnails'),
    'CHAT_ATTACHMENTS_DIR': os.path.join('data_storage', 'chat_attachments'),
    'USERS_FILE': 'users.csv',
    'CLIENTS_FILE': 'clients.csv',
    'ACTIVITY_LOG_FILE': 'activity_log.csv',
    'REVIEW_REQUESTS_FILE': 'review_requests.csv',
    'NOTIFICATIONS_FILE': 'notifications.csv',
    'PROJECT_SHARES_FILE': 'project_shares.csv',
    'TASKS_FILE': 'tasks.csv',
    'CHAT_HISTORY_FILE': 'chat_history.csv',
    'PRICE_LIST_FILE': 'Price List 2017-Rev-Edited -All Item 2018.xlsx',
    'LOCAL_PRICE_LIST_FILE': 'local_items.xlsx',
    'CHALLAN_LOG_FILE': 'challan.xlsx',
    'MARKUP': float(os.getenv('MARKUP', 0.08)),
    'BDT_CONVERSION_RATE': float(os.getenv('BDT_CONVERSION_RATE', 125.0)),
    'CUSTOMS_DUTY_PERCENTAGE': float(os.getenv('CUSTOMS_DUTY_PERCENTAGE', 0.16)), # 16%
    'MODEL_NAME': os.getenv('MODEL_NAME', 'all-MiniLM-L6-v2'),
    'ITEM_FAISS_INDEX_FILE': os.path.join('data_storage', 'item_faiss_index.bin'),
    'ITEM_SEARCH_DATA_FILE': os.path.join('data_storage', 'item_searchable_data.pkl'),
    'LOCAL_ITEM_FAISS_INDEX_FILE': os.path.join('data_storage', 'local_item_faiss_index.bin'),
    'LOCAL_ITEM_SEARCH_DATA_FILE': os.path.join('data_storage', 'local_item_searchable_data.pkl'),
    'CLIENT_FAISS_INDEX_FILE': os.path.join('data_storage', 'client_faiss_index.bin'),
    'CLIENT_SEARCH_DATA_FILE': os.path.join('data_storage', 'client_searchable_data.pkl'),
    'AI_HELPER_TOP_K': 5,
    'HEADER_COLOR_HEX': "EEE576"
}

# --- In-memory Data Storage ---
sentence_model = None
item_faiss_index, item_searchable_data = None, None
local_item_faiss_index, local_item_searchable_data = None, None
client_faiss_index, client_searchable_data = None, None
users_df, clients_df = None, None
online_users = {}

# --- Helper Functions ---
def sanitize_dirty_html(html_string):
    """
    Cleans up HTML pasted from sources like MS Word, keeping basic formatting.
    """
    if not isinstance(html_string, str) or not html_string.strip():
        return ''

    soup = BeautifulSoup(html_string, 'html.parser')

    # Whitelist of tags to keep
    allowed_tags = {'b', 'strong', 'i', 'em', 'u', 'br'}

    # Remove all tags that are not in the whitelist
    for tag in soup.find_all(True):
        if tag.name not in allowed_tags:
            tag.unwrap()  # Removes the tag but keeps its content
        else:
            tag.attrs = {}  # Strip all attributes like style, class, etc.

    # Convert aliases
    for strong_tag in soup.find_all('strong'):
        strong_tag.name = 'b'
    for em_tag in soup.find_all('em'):
        em_tag.name = 'i'

    # Convert <p> tags to <br> tags for line breaks
    for p_tag in soup.find_all('p'):
        p_tag.insert_after(soup.new_tag('br'))
        p_tag.unwrap()

    # Get the cleaned HTML as a string
    cleaned_html = str(soup)

    # Final regex clean-up for extra spaces and line breaks
    cleaned_html = cleaned_html.replace('&nbsp;', ' ')
    cleaned_html = re.sub(r'\s+', ' ', cleaned_html).strip()
    cleaned_html = re.sub(r'(<br\s*/?>\s*)+', '<br/>', cleaned_html)
    cleaned_html = re.sub(r'^(<br\s*/?>)+|(<br\s*/?>)+$', '', cleaned_html) # Trim leading/trailing breaks

    return cleaned_html

def setup_directories_and_files():
    """Sets up the necessary directories and basic CSV files if they don't exist."""
    os.makedirs(CONFIG['DATA_DIR'], exist_ok=True)
    os.makedirs(CONFIG['FOS_DIR'], exist_ok=True)
    os.makedirs(CONFIG['PROJECTS_DIR'], exist_ok=True)
    os.makedirs(CONFIG['AUTH_DIR'], exist_ok=True)
    os.makedirs(CONFIG['ASSETS_DIR'], exist_ok=True)
    os.makedirs(CONFIG['COVERS_DIR'], exist_ok=True)
    os.makedirs(CONFIG['THUMBNAILS_DIR'], exist_ok=True)
    os.makedirs(CONFIG['CHAT_ATTACHMENTS_DIR'], exist_ok=True)

    global users_df, clients_df
    users_filepath = os.path.join(CONFIG['AUTH_DIR'], CONFIG['USERS_FILE'])
    clients_filepath = os.path.join(CONFIG['DATA_DIR'], CONFIG['CLIENTS_FILE'])

    if os.path.exists(users_filepath):
        users_df = pd.read_csv(users_filepath)
    else:
        users_df = pd.DataFrame({'sl': [1, 2], 'name': ['Admin User', 'Normal User'], 'email': ['admin@example.com', 'user@example.com'], 'password': ['adminpass', 'userpass'], 'role': ['admin', 'user']})
        users_df.to_csv(users_filepath, index=False)

    if os.path.exists(clients_filepath):
        clients_df = pd.read_csv(clients_filepath)
        clients_df.columns = [c.strip() for c in clients_df.columns]
    else:
        clients_df = pd.DataFrame({'sl': [1, 2], 'client_name': ['Global Construction Ltd.', 'Modern Builders Inc.'], 'client_address': ['123 Business Bay, Dubai', '456 Skyline Ave, Abu Dhabi']})
        clients_df.to_csv(clients_filepath, index=False)

    for key in ['ACTIVITY_LOG_FILE', 'REVIEW_REQUESTS_FILE', 'NOTIFICATIONS_FILE', 'PROJECT_SHARES_FILE', 'TASKS_FILE', 'CHAT_HISTORY_FILE']:
        filepath = os.path.join(CONFIG['DATA_DIR'], CONFIG[key])
        if not os.path.exists(filepath):
            headers = {
                'ACTIVITY_LOG_FILE': ['sl', 'date', 'user_name', 'fo_name', 'file_path', 'project_id'],
                'REVIEW_REQUESTS_FILE': ['request_id', 'user_email', 'request_type', 'item_code', 'details', 'status', 'visibility', 'remarks', 'timestamp'],
                'NOTIFICATIONS_FILE': ['notification_id', 'user_email', 'message', 'is_read', 'timestamp'],
                'PROJECT_SHARES_FILE': ['share_id', 'project_id', 'owner_email', 'shared_with_email', 'permissions', 'timestamp'],
                'TASKS_FILE': ['task_id', 'assigned_by', 'assigned_to', 'task_description', 'status', 'timestamp'],
                'CHAT_HISTORY_FILE': ['message_id', 'sender_email', 'recipient_email', 'message', 'timestamp']
            }
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                csv.writer(f).writerow(headers[key])

def log_activity(user_name, fo_name, file_path, project_id):
    try:
        log_file = os.path.join(CONFIG['DATA_DIR'], CONFIG['ACTIVITY_LOG_FILE'])

        sl = 1
        if os.path.exists(log_file) and os.path.getsize(log_file) > 0:
            try:
                df = pd.read_csv(log_file)
                if not df.empty and 'sl' in df.columns:
                    last_sl = pd.to_numeric(df['sl'], errors='coerce').max()
                    if pd.notna(last_sl):
                        sl = int(last_sl) + 1
            except Exception as read_err:
                print(f"Could not read activity log to determine next sl: {read_err}")

        with open(log_file, 'a', newline='', encoding='utf-8') as f:
            csv.writer(f).writerow([sl, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), user_name, fo_name, file_path, project_id])
    except Exception as e:
        print(f"Error logging activity: {e}")

def create_notification(user_email, message):
    try:
        with open(os.path.join(CONFIG['DATA_DIR'], CONFIG['NOTIFICATIONS_FILE']), 'a', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow([str(uuid.uuid4()), user_email, message, 'no', datetime.now().isoformat()])
    except Exception as e:
        print(f"Error creating notification for {user_email}: {e}")

def find_description_column(df):
    common_names = ['description', 'desc', 'details', 'item name', 'item description', 'particulars']
    df_columns_lower = [str(col).lower().strip() for col in df.columns]

    for name in common_names:
        if name in df_columns_lower:
            original_col_index = df_columns_lower.index(name)
            return df.columns[original_col_index]

    max_avg_len, desc_col = 0, None
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]):
            continue
        try:
            avg_len = df[col].dropna().astype(str).str.len().mean()
            if avg_len > max_avg_len:
                max_avg_len, desc_col = avg_len, col
        except (TypeError, AttributeError):
            continue
    return desc_col if desc_col is not None else df.columns[0]

def check_project_permission(project_id, user_email, user_role):
    if user_role == 'admin':
        return True

    project_filepath = os.path.join(CONFIG['PROJECTS_DIR'], f"{project_id}.json")
    if os.path.exists(project_filepath):
        with open(project_filepath, 'r', encoding='utf-8') as f:
            project_data = json.load(f)
        if project_data.get('owner_email') == user_email:
            return True

    shares_filepath = os.path.join(CONFIG['DATA_DIR'], CONFIG['PROJECT_SHARES_FILE'])
    if os.path.exists(shares_filepath):
        shares_df = pd.read_csv(shares_filepath)
        is_shared = not shares_df[(shares_df['project_id'] == project_id) & (shares_df['shared_with_email'] == user_email)].empty
        if is_shared:
            return True

    return False

def safe_float(value, default=0.0):
    try:
        if value is None or value == '':
            return default
        # Attempt to convert, removing common currency symbols or commas
        if isinstance(value, str):
            value = value.replace('$', '').replace(',', '').strip()
        return float(value)
    except (ValueError, TypeError):
        return default

def update_excel_price(filepath, sheet_name, item_code, price_data):
    """Helper function to update a specific excel file."""
    try:
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            # If sheet doesn't exist, create it from the first sheet's template
            template_sheet = wb.worksheets[0]
            ws = wb.copy_worksheet(template_sheet)
            ws.title = sheet_name
        else:
            ws = wb[sheet_name]

        header = [cell.value.lower().strip() if cell.value else '' for cell in ws[1]]
        try:
            item_code_col_idx = header.index('item_code')
            price_col_idx = header.index(price_data['price_column'])
            desc_col_idx = header.index('description')
            unit_col_idx = header.index('unit')
            # Find make column, might not exist in all sheets
            make_col_idx = header.index('make') if 'make' in header else -1

        except ValueError as e:
            return (False, f"Required column not found in sheet '{sheet_name}': {e}")

        # Search for item_code to update row
        row_to_update = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[item_code_col_idx] and str(row[item_code_col_idx]).strip() == item_code:
                row_to_update = row_idx
                break

        if row_to_update:
            # Update existing row
            ws.cell(row=row_to_update, column=price_col_idx + 1, value=price_data['price_value'])
        else:
            # Add new row if item_code is not found
            new_row_values = [''] * len(header)
            new_row_values[item_code_col_idx] = item_code
            new_row_values[price_col_idx] = price_data['price_value']
            new_row_values[desc_col_idx] = price_data.get('description', '')
            new_row_values[unit_col_idx] = price_data.get('unit', 'Pcs')
            if make_col_idx != -1:
                new_row_values[make_col_idx] = price_data.get('make', 'MISC')
            ws.append(new_row_values)

        wb.save(filepath)
        return (True, f"Successfully updated item {item_code} in {os.path.basename(filepath)}.")

    except Exception as e:
        return (False, f"Error updating Excel file: {e}")

# --- Export Functions ---
def export_file(file_type, data):
    user_info = data.get('user', {})
    filename_with_ext = data.get('filename', f"FinancialOffer_NoRef.{file_type}")
    full_reference_number = data.get('referenceNumber', "FinancialOffer_NoRef")
    selected_cover = data.get('selected_cover')
    project_id = data.get('projectId')

    buffer = io.BytesIO()

    items = data.get('items', [])
    financials = data.get('financials', {})
    visible_columns = data.get('visibleColumns', {})
    summary_scopes = data.get('summaryScopes', {})

    has_freight = safe_float(financials.get('freight_foreign_usd')) > 0
    has_delivery = safe_float(financials.get('delivery_local_bdt')) > 0
    has_discount = (safe_float(financials.get('discount_foreign_usd')) > 0 or
                    safe_float(financials.get('discount_local_bdt')) > 0 or
                    safe_float(financials.get('discount_installation_bdt')) > 0)
    data['has_additional_charges'] = has_freight or has_delivery or has_discount

    has_foreign_part = visible_columns.get('foreign_price', True) and any(safe_float(item.get('foreign_total_usd')) > 0 for item in items)
    has_local_supply_part = visible_columns.get('local_supply_price') and any(safe_float(item.get('local_supply_total_bdt')) > 0 for item in items)
    has_install_part = visible_columns.get('installation_price') and any(safe_float(item.get('installation_total_bdt')) > 0 for item in items)

    data['has_foreign_part'] = has_foreign_part
    data['has_local_part'] = has_local_supply_part or has_install_part

    sub_total_usd = sum(safe_float(scope.get('total_usd')) for scope in summary_scopes.values())
    sub_total_bdt = sum(safe_float(scope.get('total_bdt')) for scope in summary_scopes.values())

    discount_bdt_total = safe_float(financials.get('discount_local_bdt')) + safe_float(financials.get('discount_installation_bdt'))

    grand_total_usd = sub_total_usd + safe_float(financials.get('freight_foreign_usd')) - safe_float(financials.get('discount_foreign_usd'))
    grand_total_bdt = sub_total_bdt + safe_float(financials.get('delivery_local_bdt')) - discount_bdt_total

    data['words_usd'] = to_words_usd(grand_total_usd) if data['has_foreign_part'] else ""
    data['words_bdt'] = to_words_bdt(grand_total_bdt) if data['has_local_part'] else ""


    if file_type == 'pdf':
        pdf_bytes = pdf_gen.generate_financial_offer_pdf(data, CONFIG['AUTH_DIR'], CONFIG['HEADER_COLOR_HEX'])

        if selected_cover:
            temp_dir = os.path.join(CONFIG['FOS_DIR'], 'temp_merge')
            os.makedirs(temp_dir, exist_ok=True)

            safe_temp_filename = re.sub(r'[\\/*?:"<>|]', "_", filename_with_ext)
            offer_pdf_path = os.path.join(temp_dir, f"temp_offer_{safe_temp_filename}.pdf")
            cover_path = os.path.join(CONFIG['COVERS_DIR'], selected_cover)
            merged_output_path = os.path.join(temp_dir, f"merged_{safe_temp_filename}.pdf")

            try:
                with open(offer_pdf_path, 'wb') as f:
                    f.write(pdf_bytes)

                if os.path.exists(cover_path):
                    print(f"Merging cover '{cover_path}' with offer '{offer_pdf_path}'")
                    cover_merger.merge_pdfs_with_resize([cover_path, offer_pdf_path], merged_output_path)

                    with open(merged_output_path, 'rb') as f:
                        buffer.write(f.read())
                else:
                    print(f"Warning: Selected cover '{selected_cover}' not found. Exporting without cover.")
                    buffer.write(pdf_bytes)

            finally:
                if os.path.exists(offer_pdf_path): os.remove(offer_pdf_path)
                if os.path.exists(merged_output_path): os.remove(merged_output_path)
        else:
            buffer.write(pdf_bytes)

        mimetype = 'application/pdf'
    elif file_type == 'xlsx':
        wb = xlsx_gen.generate_financial_offer_xlsx(data, CONFIG['AUTH_DIR'], CONFIG['HEADER_COLOR_HEX'])
        wb.save(buffer)
        mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    else:
        return "Unsupported file type", 400

    with open(os.path.join(CONFIG['FOS_DIR'], filename_with_ext), 'wb') as f:
        f.write(buffer.getvalue())

    log_name = f"[Exported {file_type.upper()}] {full_reference_number}"
    log_activity(user_info.get('name', 'Unknown'), log_name, filename_with_ext, project_id)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=filename_with_ext, mimetype=mimetype)


def generate_po_file(file_type, po_data):
    filename_with_ext = po_data.get('filename', f"PurchaseOrder_NoRef.{file_type}")
    buffer = io.BytesIO()

    if file_type == 'pdf':
        pdf_bytes = pdf_gen.generate_purchase_order_pdf(po_data, CONFIG['AUTH_DIR'], CONFIG['HEADER_COLOR_HEX'])
        buffer.write(pdf_bytes)
        mimetype = 'application/pdf'
    elif file_type == 'xlsx':
        wb = xlsx_gen.generate_purchase_order_xlsx(po_data, CONFIG['AUTH_DIR'], CONFIG['HEADER_COLOR_HEX'])
        wb.save(buffer)
        mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    else:
        return "Unsupported file type", 400

    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=filename_with_ext, mimetype=mimetype)

def find_relevant_columns(df):
    cols = { 'sl': None, 'description': None, 'quantity': None }
    df_cols_lower = {str(col).lower().strip(): col for col in df.columns}
    sl_names = ['sl', 'sl.', 'sl no', 'serial', 'serial no', '#']
    desc_names = ['description', 'desc', 'item description', 'particulars', 'item name', 'product']
    qty_names = ['qty', 'quantity', 'bal', 'balance']

    for key, names in [('sl', sl_names), ('description', desc_names), ('quantity', qty_names)]:
        for name in names:
            if name in df_cols_lower:
                cols[key] = df_cols_lower[name]
                break

    if not cols['description']:
        max_avg_len = 0
        desc_col = None
        for col in df.columns:
            if pd.api.types.is_numeric_dtype(df[col]):
                continue
            avg_len = df[col].dropna().astype(str).str.len().mean()
            if avg_len > max_avg_len:
                max_avg_len = avg_len
                desc_col = col
        cols['description'] = desc_col

    return cols

# --- API Endpoints ---
@app.route('/')
def serve_index():
    return render_template('index.html')

@app.route('/static/<path:path>')
def send_static(path):
    return send_from_directory('static', path)

@app.route('/get_tnc/<template_name>')
def get_tnc_data(template_name):
    return jsonify({'terms': tnc.get_tnc(template_name)})

@app.route('/get_cover_thumbnail/<path:pdf_filename>')
def get_cover_thumbnail(pdf_filename):
    try:
        cover_pdf_path = safe_join(CONFIG['COVERS_DIR'], pdf_filename)
        if not os.path.isfile(cover_pdf_path): return "Cover PDF not found.", 404
        thumbnail_path = cover_merger.generate_pdf_thumbnail(cover_pdf_path, CONFIG['THUMBNAILS_DIR'])
        if thumbnail_path and os.path.exists(thumbnail_path): return send_file(thumbnail_path, mimetype='image/png')
        else: return "Thumbnail could not be generated.", 500
    except Exception as e:
        print(f"Error serving thumbnail for {pdf_filename}: {e}")
        return "Error processing request.", 500

@app.route('/get_covers', methods=['GET'])
def get_covers():
    search_query = request.args.get('q', '').lower().strip()
    try:
        covers_dir = CONFIG['COVERS_DIR']
        if not os.path.exists(covers_dir):
            return jsonify([])

        all_covers = [f for f in os.listdir(covers_dir) if f.lower().endswith('.pdf')]

        if search_query:
            # REVISED: Keyword-based search logic
            search_keywords = search_query.split()
            filtered_covers = []
            for cover_name in all_covers:
                cover_name_lower = cover_name.lower()
                # Check if all keywords are present in the filename
                if all(keyword in cover_name_lower for keyword in search_keywords):
                    filtered_covers.append(cover_name)
        else:
            filtered_covers = all_covers

        return jsonify(sorted(filtered_covers))
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/upload_cover', methods=['POST'])
def upload_cover():
    if 'cover_pdf' not in request.files: return jsonify({'success': False, 'message': 'No PDF file part in the request.'}), 400
    file = request.files['cover_pdf']
    reference_number = request.form.get('referenceNumber')
    if file.filename == '': return jsonify({'success': False, 'message': 'No file selected.'}), 400
    if not reference_number: return jsonify({'success': False, 'message': 'Project reference number is required.'}), 400

    if file and file.mimetype == 'application/pdf':
        try:
            base_filename = secure_filename(reference_number)
            filename_candidate = f"{base_filename}.pdf"
            save_path = os.path.join(CONFIG['COVERS_DIR'], filename_candidate)
            counter = 2
            while os.path.exists(save_path):
                filename_candidate = f"{base_filename} ({counter}).pdf"
                save_path = os.path.join(CONFIG['COVERS_DIR'], filename_candidate)
                counter += 1
            file.save(save_path)
            cover_merger.generate_pdf_thumbnail(save_path, CONFIG['THUMBNAILS_DIR'])
            return jsonify({'success': True, 'message': 'Cover uploaded successfully.', 'savedFilename': filename_candidate})
        except Exception as e:
            print(f"Error saving uploaded cover: {e}")
            return jsonify({'success': False, 'message': f'Server error: {e}'}), 500
    else:
        return jsonify({'success': False, 'message': 'Invalid file type. Please upload a PDF.'}), 400

@app.route('/reinitialize', methods=['POST'])
def reinitialize_data_endpoint():
    if request.json.get('role') != 'admin': return jsonify({'success': False, 'message': 'Permission denied.'}), 403
    try:
        global sentence_model, item_faiss_index, item_searchable_data, local_item_faiss_index, local_item_searchable_data, client_faiss_index, client_searchable_data
        data_objects = data_management.initialize_data(CONFIG, force_rebuild=True)
        sentence_model = data_objects['sentence_model']
        item_faiss_index = data_objects['item_faiss_index']
        item_searchable_data = data_objects['item_searchable_data']
        local_item_faiss_index = data_objects['local_item_faiss_index']
        local_item_searchable_data = data_objects['local_item_searchable_data']
        client_faiss_index = data_objects['client_faiss_index']
        client_searchable_data = data_objects['client_searchable_data']
        return jsonify({'success': True, 'message': 'Data re-initialized successfully.'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Failed to re-initialize data: {e}'}), 500

@app.route('/get_activity_log', methods=['GET'])
def get_activity_log():
    if request.args.get('role') != 'admin': return jsonify({'success': False, 'message': 'Permission denied.'}), 403
    log_file_path = os.path.join(CONFIG['DATA_DIR'], CONFIG['ACTIVITY_LOG_FILE'])
    if not os.path.exists(log_file_path): return jsonify([])
    try:
        df = pd.read_csv(log_file_path)
        if df.empty: return jsonify([])
        df = df.astype(object).where(pd.notnull(df), None)
        return jsonify(df.to_dict('records'))
    except Exception as e:
        print(f"Error reading or parsing activity log file: {e}")
        return jsonify([])

@app.route('/download_fo/<path:filename>')
def download_fo(filename):
    try:
        safe_path = safe_join(CONFIG['FOS_DIR'], filename)
        if not os.path.isfile(safe_path): return "File not found.", 404
        return send_file(safe_path, as_attachment=True)
    except Exception as e:
        return "Error processing request.", 500

@app.route('/login', methods=['POST'])
def login():
    data = request.json
    users_df = pd.read_csv(os.path.join(CONFIG['AUTH_DIR'], CONFIG['USERS_FILE']))
    user = users_df[(users_df['email'] == data.get('email')) & (users_df['password'] == data.get('password'))]
    if not user.empty:
        return jsonify({'success': True, 'user': user.iloc[0].to_dict()})
    return jsonify({'success': False, 'message': 'Invalid email or password'})

@app.route('/update_user', methods=['POST'])
def update_user():
    data = request.json
    new_email, new_password = data.get('email'), data.get('password')
    try:
        user_id = int(data.get('sl'))
    except (ValueError, TypeError):
        return jsonify({'success': False, 'message': 'Invalid user ID.'}), 400
    if not all([user_id, new_email, new_password]):
        return jsonify({'success': False, 'message': 'Missing data.'}), 400
    try:
        users_filepath = os.path.join(CONFIG['AUTH_DIR'], CONFIG['USERS_FILE'])
        users_df = pd.read_csv(users_filepath)
        users_df['sl'] = users_df['sl'].astype(int)
        user_index = users_df.index[users_df['sl'] == user_id].tolist()
        if not user_index:
            return jsonify({'success': False, 'message': 'User not found.'}), 404
        users_df.loc[user_index[0], 'email'] = new_email
        users_df.loc[user_index[0], 'password'] = new_password
        users_df.to_csv(users_filepath, index=False)
        return jsonify({'success': True, 'message': 'User details updated successfully.'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'An error occurred: {e}'}), 500

@app.route('/get_sheet_names', methods=['GET'])
def get_sheet_names():
    all_sheet_names = set()
    try:
        if os.path.exists(CONFIG['PRICE_LIST_FILE']):
            wb_foreign = load_workbook(CONFIG['PRICE_LIST_FILE'], read_only=True, keep_vba=False)
            all_sheet_names.update(wb_foreign.sheetnames)

        if os.path.exists(CONFIG['LOCAL_PRICE_LIST_FILE']):
            wb_local = load_workbook(CONFIG['LOCAL_PRICE_LIST_FILE'], read_only=True, keep_vba=False)
            all_sheet_names.update(wb_local.sheetnames)

        return jsonify(sorted(list(all_sheet_names)))
    except Exception as e:
        print(f"Error reading sheet names: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/get_filter_options', methods=['GET'])
def get_filter_options():
    filter_options = {
        "make": [], "approvals": [], "model": []
    }
    filter_columns = list(filter_options.keys())

    files_to_process = [CONFIG['PRICE_LIST_FILE'], CONFIG['LOCAL_PRICE_LIST_FILE']]

    all_data_df = pd.DataFrame()

    for file_path in files_to_process:
        if os.path.exists(file_path):
            try:
                # Read all sheets from the Excel file
                xls = pd.ExcelFile(file_path)
                for sheet_name in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name=sheet_name)
                    # Normalize column names
                    df.columns = [str(c).lower().strip() for c in df.columns]
                    all_data_df = pd.concat([all_data_df, df], ignore_index=True)
            except Exception as e:
                print(f"Could not process {file_path} for filters: {e}")
                continue

    if not all_data_df.empty:
        for col in filter_columns:
            if col in all_data_df.columns:
                # Get unique non-null values, convert to string, and sort
                unique_values = all_data_df[col].dropna().astype(str).unique()
                filter_options[col] = sorted([v for v in unique_values if v.strip()])

    return jsonify(filter_options)

@app.route('/get_offer_config', methods=['GET'])
def get_offer_config():
    """Provides frontend with necessary calculation rates."""
    try:
        return jsonify({
            'success': True,
            'bdt_conversion_rate': CONFIG['BDT_CONVERSION_RATE'],
            'customs_duty_percentage': CONFIG['CUSTOMS_DUTY_PERCENTAGE']
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/search_clients', methods=['GET'])
def search_clients():
    query = request.args.get('q', '').lower()
    if not query or client_faiss_index is None or sentence_model is None: return jsonify([])
    query_embedding = sentence_model.encode([query], convert_to_tensor=True).cpu().numpy()
    _, indices = client_faiss_index.search(query_embedding, k=5)
    return jsonify([client_searchable_data[i] for i in indices[0]])

@app.route('/search_items', methods=['GET'])
def search_items():
    query = request.args.get('q', '').lower()
    role = request.args.get('role', 'user')
    source = request.args.get('source', 'all')

    make_filter = [t.strip().lower() for t in request.args.get('make', '').split(',') if t]
    approvals_filter = [t.strip().lower() for t in request.args.get('approvals', '').split(',') if t]
    model_filter = [t.strip().lower() for t in request.args.get('model', '').split(',') if t]

    # --- START MODIFICATION ---
    product_type_filter_str = request.args.get('product_type', '')
    product_type_filter = [s.strip().lower() for s in product_type_filter_str.split(',') if s] if product_type_filter_str else []
    # --- END MODIFICATION ---

    all_items_to_search = []
    if source in ['foreign', 'all'] and item_searchable_data:
        for item in item_searchable_data:
            item_copy = item.copy()
            item_copy['source_type'] = 'foreign'
            all_items_to_search.append(item_copy)

    if source in ['local', 'all'] and local_item_searchable_data:
        for item in local_item_searchable_data:
            item_copy = item.copy()
            item_copy['source_type'] = 'local'
            all_items_to_search.append(item_copy)

    # Apply categorical filters first
    filtered_items = []
    for item in all_items_to_search:
        # --- START MODIFICATION ---
        # Check sheet name
        if product_type_filter and str(item.get('product_type', '')).lower() not in product_type_filter:
            continue
        # --- END MODIFICATION ---
        # Check make
        if make_filter and str(item.get('make', '')).lower() not in make_filter:
            continue
        # Check approvals
        if approvals_filter and str(item.get('approvals', '')).lower() not in approvals_filter:
            continue
        # Check model
        if model_filter and str(item.get('model', '')).lower() not in model_filter:
            continue
        filtered_items.append(item)

    if not query:
        # If no search query, return the pre-filtered list
        return jsonify(filtered_items)

    # Parse the text query to separate positive and negative keywords
    all_terms = query.split()
    positive_query_parts = []
    negative_keywords = []

    for term in all_terms:
        if term.startswith('-') and len(term) > 1:
            negative_keywords.append(term[1:])
        else:
            positive_query_parts.append(term)

    positive_query = " ".join(positive_query_parts)
    positive_keywords = [word for word in re.split(r'[^a-z0-9]+', positive_query) if word]

    if not positive_keywords:
        return jsonify(filtered_items)

    scored_results = []
    for item in filtered_items: # Search within the already filtered items
        target_text = item.get('search_text', '')

        # 1. Negative Filter: Skip item if it contains any negative keyword
        if any(neg_keyword in target_text for neg_keyword in negative_keywords):
            continue

        # 2. Positive Scoring: Score based on positive keywords
        matched_keywords_count = sum(1 for pos_keyword in positive_keywords if pos_keyword in target_text)

        if matched_keywords_count > 0:
            score = matched_keywords_count / len(positive_keywords)

            if positive_query in target_text:
                score += 0.5

            item['relevance_score'] = score
            item['source_sort_key'] = 0 if item['source_type'] == 'foreign' else 1
            scored_results.append(item)

    # Sort results
    sorted_results = sorted(scored_results, key=lambda x: (-x['relevance_score'], x['source_sort_key']))

    if role != 'admin':
        for item in sorted_results:
            item.pop('relevance_score', None)
            item.pop('source_sort_key', None)

    return jsonify(sorted_results)

@app.route('/project', methods=['POST'])
def save_project():
    data = request.json

    # --- NEW SANITIZATION STEP ---
    items_to_save = data.get('items')
    if items_to_save:
        for item in items_to_save:
            if 'description' in item and item['description']:
                item['description'] = sanitize_dirty_html(item['description'])
        data['items'] = items_to_save
    # --- END OF SANITIZATION STEP ---

    is_new_project = not data.get('projectId')
    project_id = data.get('projectId') or str(uuid.uuid4())
    filename = f"{project_id}.json"
    filepath = os.path.join(CONFIG['PROJECTS_DIR'], filename)

    existing_status = 'Pending'
    # BUG FIX: Preserve original owner when an admin saves
    original_owner_email = data.get('user', {}).get('email') # Default to current user for new projects
    if not is_new_project and os.path.exists(filepath):
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                existing_data = json.load(f)
                existing_status = existing_data.get('status', 'Pending')
                original_owner_email = existing_data.get('owner_email', original_owner_email)
        except (IOError, json.JSONDecodeError):
            pass # If file is corrupted, proceed with current user as owner

    project_type = data.get('projectType', 'offer')

    project_data = {
        'projectId': project_id,
        'referenceNumber': data.get('referenceNumber'),
        'lastModified': datetime.now().isoformat(),
        'status': existing_status,
        'owner_email': original_owner_email, # BUG FIX: Use original owner
        'projectType': project_type,
        'items': data.get('items'),
        'client': data.get('client', {}),
    }

    if project_type in ['offer', 'po', 'challan']:
        project_data['categories'] = data.get('categories', [])

    if project_type == 'offer':
        project_data.update({
            'financials': data.get('financials', {}),
            'financialLabels': data.get('financialLabels', {}),
            'selected_cover': data.get('selected_cover'),
            'visibleColumns': data.get('visibleColumns'),
            'isSummaryPageEnabled': data.get('isSummaryPageEnabled'),
            'summaryScopeDescriptions': data.get('summaryScopeDescriptions'),
            'tncState': data.get('tncState'),
            'includeSignature': data.get('includeSignature', True)
        })
    elif project_type == 'po':
         project_data.update({
            'financials': data.get('financials', {}),
            'termsAndConditions': data.get('termsAndConditions'),
            'originalOfferRef': data.get('originalOfferRef')
        })
    elif project_type == 'challan':
         project_data.update({
            'includeSignature': data.get('includeSignature', True)
        })
    elif project_type == 'ai_helper':
        project_data['selections'] = data.get('selections')
        project_data['matchSources'] = data.get('matchSources', {})
        project_data['namingCategories'] = data.get('namingCategories', [])
        project_data['descriptionColumnIndex'] = data.get('descriptionColumnIndex', -1)
        project_data['quantityColumnIndex'] = data.get('quantityColumnIndex', -1)
        project_data['unitColumnIndex'] = data.get('unitColumnIndex', -1)
        project_data['unitPriceColumnIndex'] = data.get('unitPriceColumnIndex', -1)


    try:
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(project_data, f, indent=4)

        user_name = data.get('user', {}).get('name', 'Unknown')
        reference_number = project_data.get('referenceNumber', 'Unsaved Project')

        log_action = 'PROJECT_CREATED' if is_new_project else 'PROJECT_SAVED'
        log_activity(user_name, reference_number, log_action, project_id)

        return jsonify({'success': True, 'projectId': project_id, 'referenceNumber': project_data['referenceNumber'], 'message': 'Project saved successfully.'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

# app.py

# app.py

@app.route('/projects', methods=['GET'])
def get_projects():
    user_email = request.args.get('email')
    search_term = request.args.get('search', '').lower()
    # user_role is not needed here anymore for filtering, but kept in case other logic depends on it
    user_role = request.args.get('role')
    projects = []

    for filename in os.listdir(CONFIG['PROJECTS_DIR']):
        if not filename.endswith('.json'):
            continue
        try:
            filepath = os.path.join(CONFIG['PROJECTS_DIR'], filename)
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)

            # BUG FIX: Only show projects owned by the current user in "My Projects"
            # Admins should not see other users' projects in this specific list.
            if data.get('owner_email') != user_email:
                continue

            project_type = data.get('projectType', 'offer')
            reference_number_display = data.get('referenceNumber', 'N/A')
            client_name_display = data.get('client', {}).get('name', 'N/A')
            product_types_display = []

            if project_type == 'challan':
                client_name = data.get('client', {}).get('name', 'NOCLIENT')
                all_cats = set(i.get('make') for i in data.get('items', []))
                cats = sorted([str(c) for c in all_cats if c])
                cats_part = '_'.join(cats) if cats else 'MISC'
                abbreviation = ''.join(word[0]for word in client_name.split()).upper()
                client_part = ''.join(filter(str.isalnum, abbreviation))[:4]
                date_part = datetime.fromisoformat(data.get('lastModified')).strftime('%d-%b-%Y')
                reference_number_display = f"DC_{data.get('referenceNumber')}_{client_part}_{cats_part}_{date_part}"
                product_types_display = cats
            elif project_type == 'ai_helper':
                reference_number_display = f"[AI] {data.get('referenceNumber', 'Untitled')}"
                client_name_display = "N/A"
                product_types_display = ["AI Processed"]
            else:
                all_product_types = set(i.get('product_type', 'N/A') for i in data.get('items', []))
                product_types_display = sorted([str(m) for m in all_product_types if m and m != 'N/A'])

            if search_term and search_term not in reference_number_display.lower() and search_term not in client_name_display.lower():
                continue

            projects.append({
                'projectId': data.get('projectId'),
                'referenceNumber': reference_number_display,
                'clientName': client_name_display,
                'dateModified': data.get('lastModified'),
                'productTypes': ', '.join(product_types_display),
                'status': data.get('status', 'Pending'),
                'projectType': project_type,
                **data
            })
        except Exception as e:
            print(f"Error processing project file {filename}: {e}")
            continue

    projects.sort(key=lambda p: p['dateModified'], reverse=True)
    for i, p in enumerate(projects):
        p['sl'] = i + 1

    return jsonify(projects)

# FIX: New dedicated endpoint for the admin Activity Log to fetch ALL projects
@app.route('/all_projects_for_admin', methods=['GET'])
def get_all_projects_for_admin():
    user_email = request.args.get('email')
    search_term = request.args.get('search', '').lower()
    user_role = request.args.get('role')

    if user_role != 'admin':
        return jsonify({'success': False, 'message': 'Permission denied.'}), 403

    projects = []
    for filename in os.listdir(CONFIG['PROJECTS_DIR']):
        if not filename.endswith('.json'):
            continue
        try:
            filepath = os.path.join(CONFIG['PROJECTS_DIR'], filename)
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)

            # This endpoint intentionally does NOT filter by owner_email

            project_type = data.get('projectType', 'offer')
            reference_number_display = data.get('referenceNumber', 'N/A')
            client_name_display = data.get('client', {}).get('name', 'N/A')
            product_types_display = []

            if project_type == 'challan':
                client_name = data.get('client', {}).get('name', 'NOCLIENT')
                all_cats = set(i.get('make') for i in data.get('items', []))
                cats = sorted([str(c) for c in all_cats if c])
                cats_part = '_'.join(cats) if cats else 'MISC'
                abbreviation = ''.join(word[0]for word in client_name.split()).upper()
                client_part = ''.join(filter(str.isalnum, abbreviation))[:4]
                date_part = datetime.fromisoformat(data.get('lastModified')).strftime('%d-%b-%Y')
                reference_number_display = f"DC_{data.get('referenceNumber')}_{client_part}_{cats_part}_{date_part}"
                product_types_display = cats
            elif project_type == 'ai_helper':
                reference_number_display = f"[AI] {data.get('referenceNumber', 'Untitled')}"
                client_name_display = "N/A"
                product_types_display = ["AI Processed"]
            else:
                all_product_types = set(i.get('product_type', 'N/A') for i in data.get('items', []))
                product_types_display = sorted([str(m) for m in all_product_types if m and m != 'N/A'])

            if search_term and search_term not in reference_number_display.lower() and search_term not in client_name_display.lower():
                continue

            projects.append({
                'projectId': data.get('projectId'),
                'referenceNumber': reference_number_display,
                'clientName': client_name_display,
                'dateModified': data.get('lastModified'),
                'productTypes': ', '.join(product_types_display),
                'status': data.get('status', 'Pending'),
                'projectType': project_type,
                **data
            })
        except Exception as e:
            print(f"Error processing project file {filename}: {e}")
            continue

    projects.sort(key=lambda p: p['dateModified'], reverse=True)
    for i, p in enumerate(projects):
        p['sl'] = i + 1

    return jsonify(projects)

@app.route('/project/<project_id>', methods=['GET', 'DELETE'])
def handle_project(project_id):
    user_email = request.args.get('email')
    user_role = request.args.get('role')
    filepath = os.path.join(CONFIG['PROJECTS_DIR'], f"{project_id}.json")
    if not os.path.exists(filepath):
        return jsonify({'success': False, 'message': 'Project not found.'}), 404
    if request.method == 'GET':
        if not check_project_permission(project_id, user_email, user_role):
            return jsonify({'success': False, 'message': 'Permission denied.'}), 403
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                project_data = json.load(f)
            return jsonify({'success': True, 'data': project_data})
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500
    if request.method == 'DELETE':
        with open(filepath, 'r') as f:
            project_data = json.load(f)
        if project_data.get('status') == 'Delivered' and user_role != 'admin':
            return jsonify({'success': False, 'message': 'Only admins can delete delivered projects.'}), 403
        if not check_project_permission(project_id, user_email, user_role):
            return jsonify({'success': False, 'message': 'Permission denied.'}), 403
        try:
            os.remove(filepath)
            shares_filepath = os.path.join(CONFIG['DATA_DIR'], CONFIG['PROJECT_SHARES_FILE'])
            if os.path.exists(shares_filepath):
                shares_df = pd.read_csv(shares_filepath)
                shares_df = shares_df[shares_df['project_id'] != project_id]
                shares_df.to_csv(shares_filepath, index=False)
            return jsonify({'success': True, 'message': 'Project deleted.'})
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/project/reference/<project_id>', methods=['POST'])
def update_project_reference(project_id):
    data = request.json
    user_email = data.get('user', {}).get('email')
    user_role = data.get('user', {}).get('role')
    if not check_project_permission(project_id, user_email, user_role):
        return jsonify({'success': False, 'message': 'Permission denied.'}), 403
    new_ref = data.get('referenceNumber')
    if not new_ref:
        return jsonify({'success': False, 'message': 'New reference number is required.'}), 400
    filepath = os.path.join(CONFIG['PROJECTS_DIR'], f"{project_id}.json")
    if not os.path.exists(filepath):
        return jsonify({'success': False, 'message': 'Project not found.'}), 404
    try:
        with open(filepath, 'r+', encoding='utf-8') as f:
            project_data = json.load(f)
            project_data['referenceNumber'] = new_ref
            project_data['lastModified'] = datetime.now().isoformat()
            f.seek(0)
            json.dump(project_data, f, indent=4)
            f.truncate()
        return jsonify({'success': True, 'message': 'Reference number updated.'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/project/status/<project_id>', methods=['POST'])
def update_project_status(project_id):
    data = request.json
    user = data.get('user', {})
    user_email = user.get('email')
    user_role = user.get('role')
    if not check_project_permission(project_id, user_email, user_role):
        return jsonify({'success': False, 'message': 'Permission denied.'}), 403
    new_status = data.get('status')
    if not new_status:
        return jsonify({'success': False, 'message': 'New status is required.'}), 400
    filepath = os.path.join(CONFIG['PROJECTS_DIR'], f"{project_id}.json")
    if not os.path.exists(filepath):
        return jsonify({'success': False, 'message': 'Project not found.'}), 404
    try:
        with open(filepath, 'r+', encoding='utf-8') as f:
            project_data = json.load(f)
            project_data['status'] = new_status
            project_data['lastModified'] = datetime.now().isoformat()
            f.seek(0)
            json.dump(project_data, f, indent=4)
            f.truncate()
        if new_status == 'Delivered':
            users_df = pd.read_csv(os.path.join(CONFIG['AUTH_DIR'], CONFIG['USERS_FILE']))
            admin_emails = users_df[users_df['role'] == 'admin']['email'].tolist()
            for admin_email in admin_emails:
                create_notification(admin_email, f"Offer '{project_data.get('referenceNumber')}' has been delivered by {user.get('name')}.")
        return jsonify({'success': True, 'message': f'Project status updated to {new_status}.'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/export_pdf', methods=['POST'])
def export_pdf_endpoint():
    return export_file('pdf', request.json)

@app.route('/export_xlsx', methods=['POST'])
def export_xlsx_endpoint():
    return export_file('xlsx', request.json)

@app.route('/export_built_po', methods=['POST'])
def export_built_po_endpoint():
    data = request.json
    if data.get('user', {}).get('role') != 'admin':
        return jsonify({'success': False, 'message': 'Permission denied.'}), 403
    return generate_po_file(data.get('file_type', 'pdf'), data)

@app.route('/ai_helper/process_file', methods=['POST'])
def ai_helper_process_file():
    if 'sheet' not in request.files:
        return jsonify({'success': False, 'message': 'No file part'}), 400
    file: FileStorage = request.files['sheet']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No selected file'}), 400

    use_foreign = request.form.get('use_foreign', 'true').lower() == 'true'
    use_local = request.form.get('use_local', 'true').lower() == 'true'
    if not use_foreign and not use_local:
        return jsonify({'success': False, 'message': 'At least one item source must be selected.'}), 400

    try:
        wb = load_workbook(file, read_only=True)
        ws = wb.active

        header_row_index = -1
        header_keywords = ['sl', 'description', 'desc', 'item', 'particulars', 'qty', 'quantity', 'unit']

        for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True)):
            if row is None: continue

            row_lower = [str(cell).lower().strip() for cell in row if cell is not None]
            matches = sum(1 for keyword in header_keywords if any(keyword in cell for cell in row_lower))

            has_high_confidence_header = any(cell in ['item description', 'particulars'] for cell in row_lower)

            if matches >= 2 or has_high_confidence_header:
                header_row_index = i
                break

        if header_row_index == -1:
            raise ValueError("Could not detect a valid header row.")

        file.seek(0)
        df = pd.read_excel(file, header=header_row_index)

        df.dropna(how='all', inplace=True)
        original_headers = df.columns.tolist()
        df.columns = [str(c).strip().lower() for c in df.columns]
        lower_headers = df.columns.tolist()

        description_col_name = find_description_column(df)
        if not description_col_name:
            raise ValueError("Could not determine a description column.")

        description_col_idx = lower_headers.index(description_col_name.lower())

        qty_col_idx = -1
        unit_col_idx = -1
        unit_price_col_idx = -1
        qty_keywords = ['qty', 'quantity']
        unit_keywords = ['unit', 'units', 'uom']
        unit_price_keywords = ['price', 'rate', 'unit price', 'unit_price', 'amount']

        for i, col in enumerate(lower_headers):
            if col in qty_keywords:
                qty_col_idx = i
            if col in unit_keywords:
                unit_col_idx = i
            if col in unit_price_keywords:
                 if pd.to_numeric(df.iloc[:, i], errors='coerce').notna().any():
                    unit_price_col_idx = i


        last_valid_index = -1
        footer_keywords = ['total', 'grand total', 'subtotal', 'sub-total', 'authorized', 'signature', 'in words']

        df[description_col_name] = df[description_col_name].astype(str)

        for i in range(len(df) - 1, -1, -1):
            desc_val_series = df.iloc[i][description_col_name]
            desc_val = str(desc_val_series).lower().strip()

            if pd.notna(desc_val_series) and desc_val != 'nan' and desc_val != '' and not any(keyword in desc_val for keyword in footer_keywords):
                last_valid_index = i
                break

        if last_valid_index != -1:
            df = df.iloc[:last_valid_index + 1]

        df.dropna(subset=[description_col_name], inplace=True)
        df = df[df[description_col_name].str.strip() != '']

        processed_rows = []
        original_data_rows = df.fillna('').astype(str).values.tolist()

        for row_data in original_data_rows:
            description_text = row_data[description_col_idx]
            suggestions = []
            has_match = False

            if description_text and len(description_text.strip()) > 5:
                query_embedding = sentence_model.encode([description_text.lower()], convert_to_tensor=True).cpu().numpy()
                all_distances, all_indices = [], []

                if use_foreign and item_faiss_index:
                    dist, ind = item_faiss_index.search(query_embedding, k=CONFIG['AI_HELPER_TOP_K'])
                    all_distances.extend(dist[0])
                    all_indices.extend([(i, 'foreign') for i in ind[0] if i < len(item_searchable_data)])

                if use_local and local_item_faiss_index:
                    dist, ind = local_item_faiss_index.search(query_embedding, k=CONFIG['AI_HELPER_TOP_K'])
                    all_distances.extend(dist[0])
                    all_indices.extend([(i, 'local') for i in ind[0] if i < len(local_item_searchable_data)])

                if all_indices:
                    combined_results = sorted(zip(all_distances, all_indices), key=lambda x: x[0])
                    if combined_results and combined_results[0][0] < 1.0:
                        has_match = True

                    top_k_indices = [res[1] for res in combined_results[:CONFIG['AI_HELPER_TOP_K']]]
                    for idx, source in top_k_indices:
                        data_source = item_searchable_data if source == 'foreign' else local_item_searchable_data
                        suggestion = data_source[idx].copy()
                        suggestion['source_type'] = source
                        suggestions.append(suggestion)

            processed_rows.append({"original_data": row_data, "has_match": has_match, "suggestions": suggestions})

        return jsonify({
            "success": True,
            "headers": original_headers,
            "description_column_index": description_col_idx,
            "quantity_column_index": qty_col_idx,
            "unit_column_index": unit_col_idx,
            "unit_price_column_index": unit_price_col_idx,
            "processed_rows": processed_rows
        })

    except Exception as e:
        print(f"Intelligent parsing failed: {e}. Falling back to simple mode.")
        try:
            file.seek(0)
            df = pd.read_excel(file, header=None)
            df.dropna(how='all', inplace=True)

            if df.empty:
                 return jsonify({'success': False, 'message': 'The uploaded file appears to be empty.'}), 400

            description_col_idx = -1
            for i, col in enumerate(df.columns):
                if not df[col].isnull().all():
                    description_col_idx = i
                    break

            if description_col_idx == -1:
                return jsonify({'success': False, 'message': 'The uploaded file appears to be empty.'}), 400

            processed_rows = []
            original_data_rows = df.fillna('').astype(str).values.tolist()

            for row_data in original_data_rows:
                processed_rows.append({
                    "original_data": row_data,
                    "has_match": False,
                    "suggestions": []
                })

            return jsonify({
                "success": True,
                "headers": [f"Column {i+1}" for i in range(len(df.columns))],
                "description_column_index": description_col_idx,
                "quantity_column_index": -1,
                "unit_column_index": -1,
                "unit_price_column_index": -1,
                "processed_rows": processed_rows
            })
        except Exception as fallback_error:
            print(f"Fallback parsing also failed: {fallback_error}")
            return jsonify({'success': False, 'message': f'The file could not be processed. It might be corrupted or in an unsupported format.'}), 500


@app.route('/submit_review_request', methods=['POST'])
def submit_review_request():
    data = request.json
    try:
        with open(os.path.join(CONFIG['DATA_DIR'], CONFIG['REVIEW_REQUESTS_FILE']), 'a', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow([str(uuid.uuid4()), data.get('user_email'), data.get('request_type'), data.get('item_code'), json.dumps(data.get('details')), 'pending', data.get('visibility', 'user'), '', datetime.now().isoformat()])
        return jsonify({'success': True, 'message': 'Request submitted for your review.'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/get_my_requests', methods=['GET'])
def get_my_requests():
    user_email = request.args.get('email')
    try:
        df = pd.read_csv(os.path.join(CONFIG['DATA_DIR'], CONFIG['REVIEW_REQUESTS_FILE']))
        user_requests = df[(df['user_email'] == user_email) & (df['status'] == 'pending') & (df['visibility'] == 'user')]
        return jsonify(user_requests.to_dict('records'))
    except FileNotFoundError:
        return jsonify([])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/get_admin_requests', methods=['GET'])
def get_admin_requests():
    if request.args.get('role') != 'admin':
        return jsonify([]), 403
    try:
        df = pd.read_csv(os.path.join(CONFIG['DATA_DIR'], CONFIG['REVIEW_REQUESTS_FILE']))
        admin_requests = df[(df['status'] == 'pending') & (df['visibility'] == 'admin')]
        return jsonify(admin_requests.to_dict('records'))
    except FileNotFoundError:
        return jsonify([])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/update_review_request/<request_id>', methods=['POST'])
def update_review_request(request_id):
    data = request.json
    review_filepath = os.path.join(CONFIG['DATA_DIR'], CONFIG['REVIEW_REQUESTS_FILE'])
    df = pd.read_csv(review_filepath)
    if request_id not in df['request_id'].values:
        return jsonify({'success': False, 'message': 'Request not found.'}), 404
    request_row = df[df['request_id'] == request_id].iloc[0]
    if 'remarks' in data:
        df.loc[df['request_id'] == request_id, 'remarks'] = data['remarks']
    if 'status' in data:
        df.loc[df['request_id'] == request_id, 'status'] = data['status']
    if 'visibility' in data and data['visibility'] == 'admin':
        df.loc[df['request_id'] == request_id, 'visibility'] = 'admin'
        users_df = pd.read_csv(os.path.join(CONFIG['AUTH_DIR'], CONFIG['USERS_FILE']))
        admin_emails = users_df[users_df['role'] == 'admin']['email'].tolist()
        for admin_email in admin_emails:
            create_notification(admin_email, f"New review request from {request_row['user_email']} for item {request_row['item_code']}.")
    df.to_csv(review_filepath, index=False)
    return jsonify({'success': True, 'message': 'Request updated.'})

@app.route('/process_admin_request', methods=['POST'])
def process_admin_request():
    data = request.json
    if data.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'Permission denied.'}), 403
    request_id = data.get('request_id')
    action = data.get('action')
    requests_filepath = os.path.join(CONFIG['DATA_DIR'], CONFIG['REVIEW_REQUESTS_FILE'])
    requests_df = pd.read_csv(requests_filepath)
    request_row_series = requests_df[requests_df['request_id'] == request_id]
    if request_row_series.empty:
        return jsonify({'success': False, 'message': 'Request not found.'}), 404
    requests_df.loc[requests_df['request_id'] == request_id, 'status'] = action
    requests_df.to_csv(requests_filepath, index=False)
    request_row = request_row_series.iloc[0]
    requester_email = request_row_series.iloc[0]['user_email']
    if action == 'approved':
        create_notification(requester_email, f"Your request for item '{request_row['item_code']}' has been approved.")
        try:
            item_code = request_row['item_code']
            request_type = request_row['request_type']
            wb = load_workbook(CONFIG['PRICE_LIST_FILE'])
            item_found_and_processed = False
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                header = [cell.value for cell in ws[1]]
                try:
                    item_code_col_idx = header.index('item_code') + 1
                except ValueError:
                    continue
                for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    if str(row[item_code_col_idx - 1].value) == str(item_code):
                        if request_type == 'description_change':
                            details = json.loads(request_row['details'])
                            new_desc = details.get('new')
                            desc_col_idx = header.index('description') + 1
                            ws.cell(row=row_idx, column=desc_col_idx, value=new_desc)
                            item_found_and_processed = True
                            break
                        elif request_type == 'item_removal':
                            ws.delete_rows(row_idx, 1)
                            item_found_and_processed = True
                            break
                if item_found_and_processed:
                    break
            if not item_found_and_processed and request_type != 'item_addition':
                return jsonify({'success': False, 'message': 'Item code not found in the price list.'}), 404
            wb.save(CONFIG['PRICE_LIST_FILE'])
            if not data_management.initialize_data(CONFIG, force_rebuild=True):
                return jsonify({'success': False, 'message': 'Item processed, but failed to re-index data.'}), 500
            return jsonify({'success': True, 'message': 'Request approved and data source updated.'})
        except Exception as e:
            return jsonify({'success': False, 'message': f'Error updating Excel file: {e}'}), 500
    else:
        create_notification(requester_email, f"Your request for item '{request_row['item_code']}' was rejected.")
    return jsonify({'success': True, 'message': f'Request has been {action}.'})

@app.route('/get_notifications', methods=['GET'])
def get_notifications():
    user_email = request.args.get('email')
    try:
        df = pd.read_csv(os.path.join(CONFIG['DATA_DIR'], CONFIG['NOTIFICATIONS_FILE']))
        notifications = df[df['user_email'] == user_email].copy()
        notifications.sort_values(by='timestamp', ascending=False, inplace=True)
        return jsonify(notifications.to_dict('records'))
    except FileNotFoundError:
        return jsonify([])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/mark_notification_read', methods=['POST'])
def mark_notification_read():
    notif_id = request.json.get('notification_id')
    user_email = request.json.get('email')
    try:
        notif_filepath = os.path.join(CONFIG['DATA_DIR'], CONFIG['NOTIFICATIONS_FILE'])
        df = pd.read_csv(notif_filepath)
        df.loc[df['notification_id'] == notif_id, 'is_read'] = 'yes'
        df.to_csv(notif_filepath, index=False)

        user_notifications = df[df['user_email'] == user_email].copy()
        user_notifications.sort_values(by='timestamp', ascending=False, inplace=True)
        return jsonify({'success': True, 'notifications': user_notifications.to_dict('records')})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/delete_notification', methods=['POST'])
def delete_notification():
    notif_id = request.json.get('notification_id')
    user_email = request.json.get('email')

    if not notif_id or not user_email:
        return jsonify({'success': False, 'message': 'Missing data.'}), 400

    try:
        notif_filepath = os.path.join(CONFIG['DATA_DIR'], CONFIG['NOTIFICATIONS_FILE'])
        if not os.path.exists(notif_filepath):
             return jsonify({'success': True, 'notifications': []})

        df = pd.read_csv(notif_filepath)

        notif_exists = not df[(df['notification_id'] == notif_id) & (df['user_email'] == user_email)].empty
        if not notif_exists:
            pass
        else:
            df = df[~((df['notification_id'] == notif_id) & (df['user_email'] == user_email))]
            df.to_csv(notif_filepath, index=False)

        user_notifications = df[df['user_email'] == user_email].copy()
        user_notifications.sort_values(by='timestamp', ascending=False, inplace=True)
        return jsonify({'success': True, 'notifications': user_notifications.to_dict('records')})

    except Exception as e:
        print(f"Error deleting notification: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/share_project', methods=['POST'])
def share_project():
    data = request.json
    project_id = data.get('projectId')
    owner_email = data.get('ownerEmail')
    shared_with_email = data.get('sharedWithEmail')
    users_df = pd.read_csv(os.path.join(CONFIG['AUTH_DIR'], CONFIG['USERS_FILE']))
    if shared_with_email not in users_df['email'].values:
        return jsonify({'success': False, 'message': 'User to share with does not exist.'}), 404
    try:
        project_filepath = os.path.join(CONFIG['PROJECTS_DIR'], f"{project_id}.json")
        project_ref = "a project"
        if os.path.exists(project_filepath):
            with open(project_filepath, 'r') as f:
                project_data = json.load(f)
                project_ref = f"project '{project_data.get('referenceNumber', project_id)}'"
        with open(os.path.join(CONFIG['DATA_DIR'], CONFIG['PROJECT_SHARES_FILE']), 'a', newline='', encoding='utf-8') as f:
            csv.writer(f).writerow([str(uuid.uuid4()), project_id, owner_email, shared_with_email, 'edit', datetime.now().isoformat()])
        create_notification(shared_with_email, f"User {owner_email} has shared {project_ref} with you.")
        return jsonify({'success': True, 'message': f'Project shared with {shared_with_email}.'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/get_shared_projects', methods=['GET'])
def get_shared_projects():
    user_email = request.args.get('email')
    shared_projects = []
    try:
        shares_df = pd.read_csv(os.path.join(CONFIG['DATA_DIR'], CONFIG['PROJECT_SHARES_FILE']))
        user_shares = shares_df[shares_df['shared_with_email'] == user_email]
        for index, row in user_shares.iterrows():
            project_id = row['project_id']
            project_filepath = os.path.join(CONFIG['PROJECTS_DIR'], f"{project_id}.json")
            if os.path.exists(project_filepath):
                with open(project_filepath, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    makes = sorted(list(set(item.get('make', 'N/A') for item in data.get('items', []))))
                    shared_projects.append({
                        'sl': len(shared_projects) + 1,
                        'projectId': data.get('projectId'),
                        'referenceNumber': data.get('referenceNumber'),
                        'clientName': data.get('client', {}).get('name', 'N/A'),
                        'dateModified': data.get('lastModified'),
                        'productTypes': ', '.join(makes),
                        'status': data.get('status', 'Pending'),
                        'owner_email': data.get('owner_email'),
                        'projectType': data.get('projectType', 'offer'),
                        'share_timestamp': row['timestamp']
                    })
        shared_projects.sort(key=lambda p: p['share_timestamp'], reverse=True)
        return jsonify(shared_projects)
    except FileNotFoundError:
        return jsonify([])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/send_message', methods=['POST'])
def send_message():
    data = request.json
    sender_email, recipient_email, message = data.get('sender'), data.get('recipient'), data.get('message')
    if not all([sender_email, recipient_email, message]):
        return jsonify({'success': False, 'message': 'Missing data.'}), 400

    users_df = pd.read_csv(os.path.join(CONFIG['AUTH_DIR'], CONFIG['USERS_FILE']))
    sender_series = users_df[users_df['email'] == sender_email]

    if sender_series.empty or sender_series.iloc[0]['role'] != 'admin':
        return jsonify({'success': False, 'message': 'Permission denied. Only admins can send messages.'}), 403

    sender_name = sender_series.iloc[0]['name']

    formatted_message = f"<b>Admin Message:</b><br>{message}"

    if recipient_email.lower() == 'all':
        recipients = users_df[users_df['email'] != sender_email]['email'].tolist()
        for r_email in recipients:
            create_notification(r_email, formatted_message)
        log_activity(sender_name, "Admin Message to ALL", message[:100], "N/A")
        return jsonify({'success': True, 'message': 'Message sent to all users.'})
    else:
        if recipient_email not in users_df['email'].values:
            return jsonify({'success': False, 'message': 'Recipient user not found.'}), 404
        create_notification(recipient_email, formatted_message)
        log_activity(sender_name, f"Admin Message to {recipient_email}", message[:100], "N/A")
        return jsonify({'success': True, 'message': f'Message sent to {recipient_email}.'})


@app.route('/get_all_users', methods=['GET'])
def get_all_users():
    try:
        return jsonify(pd.read_csv(os.path.join(CONFIG['AUTH_DIR'], CONFIG['USERS_FILE'])).to_dict('records'))
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/get_new_challan_ref', methods=['GET'])
def get_new_challan_ref():
    challan_log_path = os.path.join(CONFIG['DATA_DIR'], CONFIG['CHALLAN_LOG_FILE'])
    ref_col_name = 'Ref'
    start_ref = 1001

    try:
        if not os.path.exists(challan_log_path):
            headers = ['SL', 'Ref', 'Date', 'Client', 'Description', 'Signed Copy Received', 'Remarks', 'Challan Carrier', 'Prepared by']
            df = pd.DataFrame(columns=headers)
            df.to_excel(challan_log_path, index=False, engine='openpyxl')
            return jsonify({'success': True, 'referenceNumber': start_ref})

        df = pd.read_excel(challan_log_path)
        if df.empty or ref_col_name not in df.columns:
            return jsonify({'success': True, 'referenceNumber': start_ref})

        last_ref = pd.to_numeric(df[ref_col_name], errors='coerce').max()
        if pd.isna(last_ref):
            return jsonify({'success': True, 'referenceNumber': start_ref})

        new_ref = int(last_ref) + 1
        return jsonify({'success': True, 'referenceNumber': new_ref})

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/export_challan', methods=['POST'])
def export_challan_endpoint():
    data = request.json
    file_type = data.get('fileType', 'pdf')
    ref_number = data.get('referenceNumber')
    client_info = data.get('client', {})
    user_name = data.get('user', {}).get('name', 'Unknown')
    project_id = data.get('projectId', 'N/A')

    safe_filename = data.get('filename', f"DC_{ref_number}_export.{file_type}")

    challan_log_path = os.path.join(CONFIG['DATA_DIR'], CONFIG['CHALLAN_LOG_FILE'])
    try:
        df = pd.read_excel(challan_log_path)
        new_sl = (df['SL'].max() + 1) if not df.empty and pd.notna(df['SL'].max()) else 1

        description = ', '.join(data.get('categories', []))

        new_row_data = {
            'SL': new_sl,
            'Ref': ref_number,
            'Date': datetime.now().strftime('%d-%b-%Y'),
            'Client': client_info.get('name', 'N/A'),
            'Description': description,
            'Signed Copy Received': '',
            'Remarks': 'Not processed yet',
            'Challan Carrier': '',
            'Prepared by': user_name
        }

        new_row_df = pd.DataFrame([new_row_data])
        df = pd.concat([df, new_row_df], ignore_index=True)

        df.to_excel(challan_log_path, index=False, engine='openpyxl')

    except Exception as e:
        print(f"Could not update challan log: {e}")

    buffer = io.BytesIO()

    if file_type == 'pdf':
        pdf_bytes = pdf_gen.generate_challan_pdf(data, CONFIG['AUTH_DIR'], CONFIG['HEADER_COLOR_HEX'])
        buffer.write(pdf_bytes)
        mimetype = 'application/pdf'
    elif file_type == 'xlsx':
        wb = xlsx_gen.generate_challan_xlsx(data, CONFIG['AUTH_DIR'], CONFIG['HEADER_COLOR_HEX'])
        wb.save(buffer)
        mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    else:
        return jsonify({'success': False, 'message': "Unsupported file type"}), 400

    log_name = f"[Challan Exported {file_type.upper()}] {ref_number}"
    log_activity(user_name, log_name, safe_filename, project_id)

    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=safe_filename, mimetype=mimetype)

@app.route('/convert_to_words', methods=['POST'])
def convert_to_words_endpoint():
    data = request.json
    usd_value = data.get('usd_value', 0)
    bdt_value = data.get('bdt_value', 0)

    words_usd = to_words_usd(usd_value)
    words_bdt = to_words_bdt(bdt_value)

    return jsonify({
        'success': True,
        'words_usd': words_usd,
        'words_bdt': words_bdt
    })

@app.route('/update_master_price', methods=['POST'])
def update_master_price():
    data = request.json
    admin_email = data.get('adminEmail')

    users_df = pd.read_csv(os.path.join(CONFIG['AUTH_DIR'], CONFIG['USERS_FILE']))
    user_role = users_df[users_df['email'] == admin_email]['role'].iloc[0]
    if user_role != 'admin':
        return jsonify({'success': False, 'message': 'Permission denied.'}), 403

    item_code = data.get('itemCode')
    price_type = data.get('priceType')
    price_value = safe_float(data.get('priceValue'))
    source_type = data.get('sourceType')
    make = data.get('productType', 'MISC')

    if not item_code:
        return jsonify({'success': False, 'message': 'Item Code is required to update the master list.'}), 400

    value_to_save = price_value
    price_column_to_update = None

    if price_type == 'installation_price':
        price_column_to_update = 'installation'
    elif price_type in ['foreign_price', 'local_supply_price']:
        markup = CONFIG.get('MARKUP', 0.08)
        value_to_save = price_value / (1 + markup)
        price_column_to_update = 'po_price'
    elif price_type == 'po_price':
        price_column_to_update = 'po_price'

    if not price_column_to_update:
        return jsonify({'success': False, 'message': f'Invalid or unhandled price type: {price_type}'}), 400

    if source_type == 'foreign':
        filepath = CONFIG['PRICE_LIST_FILE']
    else: # local
        filepath = CONFIG['LOCAL_PRICE_LIST_FILE']

    price_data = {
        'price_column': price_column_to_update,
        'price_value': value_to_save,
        'description': data.get('description'),
        'unit': data.get('unit'),
        'make': make
    }

    success, message = update_excel_price(filepath, make, item_code, price_data)

    if success:
        log_activity(admin_email, "Master Price Update", f"Updated {item_code} to {value_to_save}", "N/A")
        # print("Re-initializing data after price update...")
        # data_management.initialize_data(CONFIG, force_rebuild=True)
        # print("Data re-initialized.")

    return jsonify({'success': success, 'message': message})


@app.route('/autofill_master_prices', methods=['POST'])
def autofill_master_prices():
    admin_email = request.json.get('adminEmail')
    users_df = pd.read_csv(os.path.join(CONFIG['AUTH_DIR'], CONFIG['USERS_FILE']))
    user_role = users_df[users_df['email'] == admin_email]['role'].iloc[0]
    if user_role != 'admin':
        return jsonify({'success': False, 'message': 'Permission denied.'}), 403

    files_to_process = {
        'foreign': CONFIG['PRICE_LIST_FILE'],
        'local': CONFIG['LOCAL_PRICE_LIST_FILE']
    }

    updated_items_count = 0
    errors = []

    for source, filepath in files_to_process.items():
        try:
            if not os.path.exists(filepath):
                continue

            wb = load_workbook(filepath)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                header = [cell.value.lower().strip() if cell.value else '' for cell in ws[1]]

                try:
                    offer_price_col_idx = header.index('offer_price')
                    if 'installation' in header:
                        install_price_col_idx = header.index('installation')
                    else:
                        # If 'installation' column doesn't exist, add it
                        install_price_col_idx = len(header)
                        ws.cell(row=1, column=install_price_col_idx + 1, value='installation')
                except ValueError as e:
                    errors.append(f"Skipping sheet '{sheet_name}' in {os.path.basename(filepath)}: Missing required column ({e})")
                    continue

                for row_idx in range(2, ws.max_row + 1):
                    install_cell = ws.cell(row=row_idx, column=install_price_col_idx + 1)
                    install_price = safe_float(install_cell.value)

                    if install_price == 0:
                        offer_price_cell = ws.cell(row=row_idx, column=offer_price_col_idx + 1)
                        offer_price = safe_float(offer_price_cell.value)

                        if offer_price > 0:
                            # Logic: Installation is 10% of offer price, rounded to nearest whole number
                            calculated_install_price = round(offer_price * 0.10)
                            install_cell.value = calculated_install_price
                            updated_items_count += 1

            wb.save(filepath)

        except Exception as e:
            errors.append(f"Failed to process {os.path.basename(filepath)}: {e}")

    if updated_items_count > 0:
        log_activity(admin_email, "Master Price Autofill", f"Auto-filled {updated_items_count} items.", "N/A")
        print("Re-initializing data after autofill...")
        data_management.initialize_data(CONFIG, force_rebuild=True)
        print("Data re-initialized.")
        message = f"Successfully updated {updated_items_count} item(s). "
    else:
        message = "No items needed updating. "

    if errors:
        message += "Encountered errors: " + "; ".join(errors)
        return jsonify({'success': False, 'message': message})

    return jsonify({'success': True, 'message': message})

# --- SocketIO Chat Handlers ---
@app.route('/upload_chat_attachment', methods=['POST'])
def upload_chat_attachment():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No selected file'}), 400

    try:
        original_filename = secure_filename(file.filename)
        file_ext = os.path.splitext(original_filename)[1]
        unique_filename = f"{uuid.uuid4()}{file_ext}"
        save_path = os.path.join(CONFIG['CHAT_ATTACHMENTS_DIR'], unique_filename)
        file.save(save_path)

        file_url = f"/chat_attachment/{unique_filename}"

        return jsonify({
            'success': True,
            'url': file_url,
            'filename': original_filename
        })
    except Exception as e:
        print(f"Error uploading chat file: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/chat_attachment/<path:filename>')
def get_chat_attachment(filename):
    try:
        original_name = request.args.get('original_name', filename)
        return send_from_directory(
            CONFIG['CHAT_ATTACHMENTS_DIR'],
            filename,
            as_attachment=True,
            download_name=original_name
        )
    except FileNotFoundError:
        return "File not found.", 404

@socketio.on('connect')
def handle_connect():
    print(f"Client connected: {request.sid}")

@socketio.on('user_online')
def handle_user_online(data):
    email = data.get('email')
    if email:
        online_users[email] = request.sid
        print(f"User online: {email} with sid {request.sid}")

        users_df_path = os.path.join(CONFIG['AUTH_DIR'], CONFIG['USERS_FILE'])
        if os.path.exists(users_df_path):
            users_df = pd.read_csv(users_df_path)
            all_users_list = users_df.to_dict('records')
            online_user_emails = list(online_users.keys())
            emit('update_user_list', {'all_users': all_users_list, 'online_users': online_user_emails}, broadcast=True)

@socketio.on('disconnect')
def handle_disconnect():
    disconnected_email = None
    for email, sid in list(online_users.items()):
        if sid == request.sid:
            disconnected_email = email
            break
    if disconnected_email:
        del online_users[disconnected_email]
        print(f"User disconnected: {disconnected_email}")

        users_df_path = os.path.join(CONFIG['AUTH_DIR'], CONFIG['USERS_FILE'])
        if os.path.exists(users_df_path):
            users_df = pd.read_csv(users_df_path)
            all_users_list = users_df.to_dict('records')
            online_user_emails = list(online_users.keys())
            emit('update_user_list', {'all_users': all_users_list, 'online_users': online_user_emails}, broadcast=True)

@socketio.on('private_message')
def handle_private_message(data):
    recipient_email = data.get('recipient_email')
    message_obj = data.get('message')

    sender_email = None
    for email, sid in online_users.items():
        if sid == request.sid:
            sender_email = email
            break

    if sender_email and recipient_email and message_obj:
        timestamp = datetime.now().isoformat()
        message_id = str(uuid.uuid4())

        try:
            with open(os.path.join(CONFIG['DATA_DIR'], CONFIG['CHAT_HISTORY_FILE']), 'a', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow([message_id, sender_email, recipient_email, json.dumps(message_obj), timestamp])
        except Exception as e:
            print(f"Error saving chat message: {e}")

        recipient_sid = online_users.get(recipient_email)

        message_payload = {
            'sender_email': sender_email,
            'message': message_obj,
            'timestamp': timestamp
        }

        if recipient_sid:
            emit('new_message', message_payload, to=recipient_sid)

        emit('new_message', {**message_payload, 'recipient_email': recipient_email}, to=request.sid)

@app.route('/chat_history/<user1_email>/<user2_email>')
def get_chat_history(user1_email, user2_email):
    history = []
    try:
        chat_history_path = os.path.join(CONFIG['DATA_DIR'], CONFIG['CHAT_HISTORY_FILE'])
        if os.path.exists(chat_history_path):
            df = pd.read_csv(chat_history_path)
            conversation = df[
                ((df['sender_email'] == user1_email) & (df['recipient_email'] == user2_email)) |
                ((df['sender_email'] == user2_email) & (df['recipient_email'] == user1_email))
            ].copy()

            def parse_message(msg):
                try:
                    return json.loads(msg)
                except (json.JSONDecodeError, TypeError):
                    return {'type': 'text', 'content': msg}

            conversation['message'] = conversation['message'].apply(parse_message)
            history = conversation.sort_values(by='timestamp').to_dict('records')

    except Exception as e:
        print(f"Error fetching chat history: {e}")
        return jsonify({'success': False, 'message': 'Could not fetch history'}), 500
    return jsonify({'success': True, 'history': history})


if __name__ == '__main__':
    try:
        setup_directories_and_files()

        print("Initializing data management module...")
        init_config = CONFIG.copy()
        init_config['USERS_FILE'] = os.path.join(CONFIG['AUTH_DIR'], CONFIG['USERS_FILE'])
        init_config['CLIENTS_FILE'] = os.path.join(CONFIG['DATA_DIR'], CONFIG['CLIENTS_FILE'])
        init_config['PRICE_LIST_FILE'] = os.path.join(CONFIG['DATA_DIR'], CONFIG['PRICE_LIST_FILE'])
        init_config['LOCAL_PRICE_LIST_FILE'] = os.path.join(CONFIG['DATA_DIR'], CONFIG['LOCAL_PRICE_LIST_FILE'])

        data_objects = data_management.initialize_data(init_config)

        sentence_model = data_objects['sentence_model']
        item_faiss_index = data_objects['item_faiss_index']
        item_searchable_data = data_objects['item_searchable_data']
        local_item_faiss_index = data_objects['local_item_faiss_index']
        local_item_searchable_data = data_objects['local_item_searchable_data']
        client_faiss_index = data_objects['client_faiss_index']
        client_searchable_data = data_objects['client_searchable_data']

        print("Application initialized successfully!")
        socketio.run(app, host='0.0.0.0', debug=True, use_reloader=False, port=5001)

    except Exception as e:
        print(f"FATAL: Failed to initialize application. Please check the errors above. Exception: {e}")